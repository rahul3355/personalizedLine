"""Utilities for streaming Supabase storage files without loading them entirely into memory."""
from __future__ import annotations

import csv
import os
import tempfile
from typing import List

import httpx
from openpyxl import load_workbook


class FileStreamingError(RuntimeError):
    """Raised when a Supabase file cannot be streamed locally."""


async def stream_input_to_tempfile(
    supabase_client,
    file_path: str,
    *,
    expires_in: int = 60,
    chunk_size: int = 1024 * 1024,
) -> str:
    """Download a Supabase input file to a temporary location via streaming.

    Parameters
    ----------
    supabase_client:
        Initialized Supabase client used to create signed URLs.
    file_path:
        Path inside the ``inputs`` bucket.
    expires_in:
        Seconds the signed URL remains valid.
    chunk_size:
        Number of bytes pulled per streaming iteration.

    Returns
    -------
    str
        Filesystem path to the streamed temporary file. The caller is
        responsible for deleting the file.
    """

    try:
        signed = supabase_client.storage.from_("inputs").create_signed_url(
            file_path, expires_in
        )
    except Exception as exc:  # pragma: no cover - network/SDK errors
        raise FileStreamingError(f"Signed URL error: {exc}") from exc

    signed_url = (signed or {}).get("signedURL")
    if not signed_url:
        raise FileStreamingError("Failed to obtain signed URL for streaming")

    tmp_file = tempfile.NamedTemporaryFile(delete=False)
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            async with client.stream("GET", signed_url) as response:
                if response.status_code != 200:
                    raise FileStreamingError(
                        f"Download failed with status {response.status_code}"
                    )
                async for chunk in response.aiter_bytes(chunk_size):
                    if chunk:
                        tmp_file.write(chunk)
        tmp_file.flush()
        tmp_file.close()
        return tmp_file.name
    except Exception:
        try:
            tmp_file.close()
        finally:
            try:
                os.unlink(tmp_file.name)
            except OSError:
                pass
        raise


def extract_csv_headers(path: str) -> List[str]:
    """Return the header row from a streamed CSV file."""

    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
    return [value if value is not None else "" for value in header]


def count_csv_rows(path: str) -> int:
    """Count data rows in a streamed CSV file without loading everything."""

    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        # Skip header row to mirror pandas default behaviour
        next(reader, None)
        count = sum(1 for _ in reader)
    return count


def extract_xlsx_headers(path: str) -> List[str]:
    """Return the header row from a streamed XLSX file using read-only mode."""

    workbook = load_workbook(path, read_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(values_only=True), None) or []
        return ["" if cell is None else str(cell) for cell in header_row]
    finally:
        workbook.close()


def count_xlsx_rows(path: str) -> int:
    """Count data rows in a streamed XLSX file without materializing the sheet."""

    workbook = load_workbook(path, read_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        # Skip header row to mirror pandas' behaviour
        next(rows_iter, None)
        count = 0
        for _ in rows_iter:
            count += 1
        return count
    finally:
        workbook.close()
