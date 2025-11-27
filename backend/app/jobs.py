import os
import csv
import json
import time
import math
from typing import Optional, List, Dict, Tuple
import pandas as pd
import traceback
import tempfile
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from backend.app.gpt_helpers import generate_full_email_body
from backend.app.research import perform_research
from backend.app.email_cleaning import clean_email_body
from backend.app.supabase_client import supabase
from datetime import datetime, timedelta
import redis
import rq
import requests
from redis.exceptions import LockError
from rq import get_current_job
from supabase import StorageException
from openpyxl import load_workbook

# -----------------------------
# Redis connection
# -----------------------------
redis_conn = redis.Redis(host="redis", port=6379, decode_responses=True)
queue = rq.Queue("default", connection=redis_conn)

# Job notification channel
JOB_NOTIFICATION_CHANNEL = "job_notifications"


def publish_job_notification(job_id: str):
    """
    Publish a job notification to Redis pub/sub for instant worker pickup.

    This eliminates the polling delay, allowing workers to start processing
    jobs immediately when they are enqueued.
    """
    try:
        redis_conn.publish(JOB_NOTIFICATION_CHANNEL, job_id)
        print(f"[Jobs] Published notification for job {job_id}")
    except Exception as exc:
        # Non-critical: workers will still pick up jobs via fallback polling
        print(f"[Jobs] Warning: Failed to publish job notification: {exc}")


RAW_CHUNK_BASE_DIR = "/data/raw_chunks"
RAW_CHUNK_BUCKET = "inputs"

# Parallel processing configuration
PARALLEL_ROWS_PER_WORKER = int(os.getenv('PARALLEL_ROWS_PER_WORKER', '20'))

GENERATED_OUTPUT_COLUMNS = ("email_body", "sif_personalized_line")


def _ensure_dict(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


def _parse_output_column_filter(meta: Optional[dict]) -> Optional[List[str]]:
    meta = _ensure_dict(meta)
    columns = meta.get("output_columns")
    if not columns:
        return None
    if isinstance(columns, str):
        try:
            parsed = json.loads(columns)
        except Exception:
            return None
        if isinstance(parsed, list):
            columns = parsed
        else:
            return None
    if isinstance(columns, list):
        filtered = [str(col).strip() for col in columns if isinstance(col, (str, int))]
        filtered = [col for col in filtered if col]
        return filtered or None
    return None


def _resolve_output_header_order(
    input_headers: List[str], meta: Optional[dict]
) -> Tuple[List[str], Optional[str], List[str], List[str]]:
    meta = _ensure_dict(meta)
    allowed_columns = _parse_output_column_filter(meta)
    email_header = (meta.get("email_col") or "").strip()

    base_headers = []
    for header in input_headers:
        if header in GENERATED_OUTPUT_COLUMNS:
            continue
        if not header:
            continue
        base_headers.append(header)

    if allowed_columns:
        allowed_set = set(allowed_columns)
        base_headers = [header for header in base_headers if header in allowed_set]

    row_headers = base_headers
    if email_header:
        row_headers = [header for header in base_headers if header != email_header]

    final_headers = list(row_headers)
    if email_header:
        final_headers.append(email_header)
    for generated in GENERATED_OUTPUT_COLUMNS:
        if generated not in final_headers:
            final_headers.append(generated)

    if allowed_columns:
        persist_headers = list(row_headers)
        if email_header and email_header not in persist_headers:
            persist_headers.append(email_header)
    else:
        persist_headers = list(base_headers)
        if email_header and email_header not in persist_headers:
            persist_headers.append(email_header)

    return row_headers, email_header or None, final_headers, persist_headers


def _format_storage_error(error) -> str:
    if isinstance(error, dict):
        message = error.get("message")
        if message:
            return message
        return json.dumps(error)
    return str(error)


def _upload_to_storage(storage_path: str, file_obj, context: str, bucket: str = "outputs"):
    """Upload a file to Supabase storage and raise if the response indicates failure."""
    try:
        response = supabase.storage.from_(bucket).upload(storage_path, file_obj)
    except StorageException as exc:
        raise RuntimeError(
            f"Supabase upload failed for {context} ({storage_path}): {exc}"
        ) from exc

    if response is None:
        raise RuntimeError(
            f"Supabase upload failed for {context} ({storage_path}): empty response"
        )

    status_code = getattr(response, "status_code", None)
    if status_code and status_code >= 400:
        raise RuntimeError(
            f"Supabase upload failed for {context} ({storage_path}): HTTP {status_code}"
        )

    try:
        payload = response.json()
    except Exception:
        payload = None

    if isinstance(payload, dict):
        error = payload.get("error")
        if error:
            raise RuntimeError(
                f"Supabase upload failed for {context} ({storage_path}): {_format_storage_error(error)}"
            )
        if payload.get("message") and payload.get("statusCode"):
            raise RuntimeError(
                f"Supabase upload failed for {context} ({storage_path}): {payload['statusCode']}: {payload['message']}"
            )

    return response


def _remove_from_storage(storage_path: str, context: str, bucket: str = "outputs"):
    try:
        supabase.storage.from_(bucket).remove([storage_path])
    except StorageException as exc:
        print(f"[Worker] Warning: failed to remove {context} ({storage_path}) from {bucket}: {exc}")
    except Exception as exc:
        print(f"[Worker] Warning: unexpected error removing {context} ({storage_path}) from {bucket}: {exc}")


def _chunk_raw_local_path(job_id: str, chunk_id: int) -> str:
    local_dir = os.path.join(RAW_CHUNK_BASE_DIR, job_id)
    os.makedirs(local_dir, exist_ok=True)
    return os.path.join(local_dir, f"chunk_{chunk_id}.csv")


def _persist_chunk_rows(job_id: str, chunk_id: int, headers: List[str], rows: List[Dict[str, str]], user_id: str) -> str:
    headers = list(headers or [])
    local_path = _chunk_raw_local_path(job_id, chunk_id)
    with open(local_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            normalized = {}
            for header in headers:
                value = row.get(header, "")
                normalized[header] = "" if value is None else value
            writer.writerow(normalized)

    storage_path = f"{user_id}/{job_id}/raw_chunks/chunk_{chunk_id}.csv"
    with open(local_path, "rb") as f:
        _upload_to_storage(
            storage_path,
            f,
            f"raw chunk {chunk_id} for job {job_id}",
            bucket=RAW_CHUNK_BUCKET,
        )

    return storage_path


def _download_chunk_from_storage(storage_path: str) -> str:
    data = supabase.storage.from_(RAW_CHUNK_BUCKET).download(storage_path)
    if data is None:
        raise RuntimeError(f"Empty response downloading chunk {storage_path} from storage")

    if isinstance(data, bytes):
        payload = data
    elif hasattr(data, "read"):
        payload = data.read()
    elif isinstance(data, str):
        payload = data.encode("utf-8")
    else:
        raise RuntimeError(f"Unsupported download payload type for chunk {storage_path}: {type(data)}")

    tmp_dir = tempfile.mkdtemp(prefix="chunk_raw_")
    local_path = os.path.join(tmp_dir, os.path.basename(storage_path))
    with open(local_path, "wb") as f:
        f.write(payload)

    return local_path


def _csv_headers_and_total(path: str):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, None) or []
        total = sum(1 for _ in reader)
    return headers, total


def _csv_headers_only(path: str):
    """Get CSV headers without counting rows (fast)."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, None) or []
    return headers


def _iter_csv_rows(path: str):
    def generator():
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield {key: (value if value is not None else "") for key, value in row.items()}

    return generator()


def _xlsx_headers_and_total(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None) or []
        headers = [str(cell) if cell is not None else "" for cell in headers_row]
        max_row = ws.max_row or 0
        total = max(max_row - 1, 0) if headers else 0
        return headers, total
    finally:
        wb.close()


def _xlsx_headers_only(path: str):
    """Get XLSX headers without counting rows (fast)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None) or []
        headers = [str(cell) if cell is not None else "" for cell in headers_row]
        return headers
    finally:
        wb.close()


def _iter_xlsx_rows(path: str, headers: List[str]):
    def generator():
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            # Skip header row if present
            next(rows_iter, None)
            for row in rows_iter:
                values = list(row)
                row_dict = {}
                for idx, column in enumerate(headers):
                    value = values[idx] if idx < len(values) else None
                    row_dict[column] = "" if value is None else value
                yield row_dict
        finally:
            wb.close()

    return generator()


def _input_iterator(local_path: str):
    ext = os.path.splitext(local_path)[1].lower()
    if ext == ".csv":
        headers, total = _csv_headers_and_total(local_path)
        return headers, total, _iter_csv_rows(local_path)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        headers, total = _xlsx_headers_and_total(local_path)
        return headers, total, _iter_xlsx_rows(local_path, headers)
    raise RuntimeError(f"Unsupported file type: {ext}")


def _get_headers_and_iterator(local_path: str, cached_total: int):
    """
    Get headers and row iterator without counting rows (uses cached total).
    This is significantly faster than _input_iterator for large files.
    """
    ext = os.path.splitext(local_path)[1].lower()
    if ext == ".csv":
        headers = _csv_headers_only(local_path)
        return headers, cached_total, _iter_csv_rows(local_path)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        headers = _xlsx_headers_only(local_path)
        return headers, cached_total, _iter_xlsx_rows(local_path, headers)
    raise RuntimeError(f"Unsupported file type: {ext}")


def _finalize_empty_job(
    job_id: str,
    user_id: str,
    final_headers: Optional[List[str]],
    timings: dict,
    job_start: float,
):
    columns: List[str] = []
    seen = set()

    for header in list(final_headers or []):
        if not header or header in seen:
            continue
        columns.append(header)
        seen.add(header)

    for generated in GENERATED_OUTPUT_COLUMNS:
        if generated not in seen:
            columns.append(generated)
            seen.add(generated)

    if not columns:
        columns = list(GENERATED_OUTPUT_COLUMNS)

    empty_df = pd.DataFrame(columns=columns)
    local_dir = tempfile.mkdtemp()
    out_csv = os.path.join(local_dir, f"{job_id}_final.csv")
    out_xlsx = os.path.join(local_dir, f"{job_id}_final.xlsx")
    try:
        empty_df.to_csv(out_csv, index=False)
        empty_df.to_excel(out_xlsx, index=False, engine="openpyxl")

        storage_path = f"{user_id}/{job_id}/result.xlsx"
        with open(out_xlsx, "rb") as f:
            _upload_to_storage(storage_path, f, f"final result for job {job_id}")

        timings["process_job_total"] = record_time("process_job total", job_start, job_id)
        supabase.table("jobs").update(
            {
                "status": "succeeded",
                "finished_at": datetime.utcnow().isoformat() + "Z",
                "result_path": storage_path,
                "progress_percent": 100,
                "timing_json": json.dumps(timings),
            }
        ).eq("id", job_id).execute()
    finally:
        import shutil

        shutil.rmtree(local_dir, ignore_errors=True)

    print(f"[Worker] Job {job_id} contained no rows; generated empty result file")


def refund_job_credits(job_id: str, user_id: Optional[str], reason: str = "") -> bool:
    """
    Refund credits for a job if they were previously deducted.
    Refunds to the same buckets (monthly/addon) from which they were deducted.
    """
    try:
        job_res = (
            supabase.table("jobs")
            .select("meta_json, user_id")
            .eq("id", job_id)
            .limit(1)
            .execute()
        )
        if not job_res.data:
            return False

        job_row = job_res.data[0]
        meta = _ensure_dict(job_row.get("meta_json"))
        user_id = user_id or job_row.get("user_id")
        if not user_id:
            return False
        cost = meta.get("credit_cost")
        if not cost or cost <= 0:
            return False
        if not meta.get("credits_deducted"):
            return False
        if meta.get("credits_refunded"):
            return False

        # Get breakdown of how credits were deducted
        monthly_deducted = meta.get("monthly_deducted", 0)
        addon_deducted = meta.get("addon_deducted", 0)

        # If no breakdown available (legacy jobs), refund all to monthly bucket
        if monthly_deducted == 0 and addon_deducted == 0:
            monthly_deducted = cost
            addon_deducted = 0

        max_attempts = 5
        for _ in range(max_attempts):
            profile_res = (
                supabase.table("profiles")
                .select("credits_remaining, addon_credits")
                .eq("id", user_id)
                .limit(1)
                .execute()
            )

            if not profile_res.data:
                print(f"[Credits] Profile not found for user {user_id}")
                return False

            profile_data = profile_res.data[0]
            current_monthly = int(profile_data.get("credits_remaining") or 0)
            current_addon = int(profile_data.get("addon_credits") or 0)

            # Refund to the same buckets from which credits were deducted
            new_monthly = current_monthly + monthly_deducted
            new_addon = current_addon + addon_deducted

            update_res = (
                supabase.table("profiles")
                .update({
                    "credits_remaining": new_monthly,
                    "addon_credits": new_addon
                })
                .eq("id", user_id)
                .eq("credits_remaining", current_monthly)
                .eq("addon_credits", current_addon)
                .execute()
            )

            if getattr(update_res, "error", None):
                raise RuntimeError(getattr(update_res, "error", "Failed to refund credits"))

            updated_rows = getattr(update_res, "data", None) or []
            if updated_rows:
                print(
                    f"[Credits] Refunded {cost} credits for job {job_id} "
                    f"(monthly: +{monthly_deducted}, addon: +{addon_deducted}) - {reason}"
                )
                break

            time.sleep(0.1)
        else:
            raise RuntimeError("Failed to atomically refund credits")

        # Record in ledger
        supabase.table("ledger").insert(
            {
                "user_id": user_id,
                "change": cost,
                "amount": 0.0,
                "reason": f"job refund: {job_id}{' - ' + reason if reason else ''}",
                "ts": datetime.utcnow().isoformat(),
            }
        ).execute()

        meta["credits_refunded"] = True
        supabase.table("jobs").update({"meta_json": meta}).eq("id", job_id).execute()

        return True
    except Exception as exc:
        print(f"[Credits] Failed to refund credits for job {job_id}: {exc}")
        return False

def _deduct_job_credits(
    job_id: str,
    user_id: str,
    total: int,
    meta: dict,
    *,
    supabase_client=supabase,
) -> bool:
    """
    Deduct credits for a job if they have not already been charged.
    Uses two-bucket system: deducts from monthly credits first, then add-on credits.
    """

    if total <= 0:
        return True

    if meta.get("credits_deducted"):
        print(
            f"[Credits] Job {job_id} already has credits deducted; skipping new deduction"
        )
        return True

    deducted = False
    previous_monthly = 0
    previous_addon = 0
    monthly_deducted = 0
    addon_deducted = 0

    try:
        max_attempts = 5
        for _ in range(max_attempts):
            # Read both credit buckets
            profile_res = (
                supabase_client.table("profiles")
                .select("credits_remaining, addon_credits")
                .eq("id", user_id)
                .limit(1)
                .execute()
            )

            if not profile_res.data:
                print(f"[Credits] Profile not found for user {user_id}")
                return False

            profile_data = profile_res.data[0]
            previous_monthly = int(profile_data.get("credits_remaining") or 0)
            previous_addon = int(profile_data.get("addon_credits") or 0)
            total_available = previous_monthly + previous_addon

            # Check if sufficient credits
            if total_available < total:
                supabase_client.table("jobs").update(
                    {
                        "status": "failed",
                        "finished_at": datetime.utcnow().isoformat() + "Z",
                        "error": "Not enough credits",
                    }
                ).eq("id", job_id).execute()
                print(
                    f"[Credits] Job {job_id} has insufficient credits "
                    f"(monthly={previous_monthly}, addon={previous_addon}, "
                    f"total={total_available} < needed={total})"
                )
                return False

            # Two-bucket deduction: use monthly credits first, then add-on credits
            if previous_monthly >= total:
                # Use only monthly credits
                new_monthly = previous_monthly - total
                new_addon = previous_addon
                monthly_deducted = total
                addon_deducted = 0
                deduction_source = "monthly"
            else:
                # Use all monthly + some add-on credits
                remaining_needed = total - previous_monthly
                new_monthly = 0
                new_addon = previous_addon - remaining_needed
                monthly_deducted = previous_monthly
                addon_deducted = remaining_needed
                deduction_source = "monthly+addon"

            # Atomic update with CAS on both columns
            update_res = (
                supabase_client.table("profiles")
                .update({
                    "credits_remaining": new_monthly,
                    "addon_credits": new_addon
                })
                .eq("id", user_id)
                .eq("credits_remaining", previous_monthly)
                .eq("addon_credits", previous_addon)
                .execute()
            )

            if getattr(update_res, "error", None):
                raise RuntimeError(
                    getattr(update_res, "error", "Failed to update credits")
                )

            updated_rows = getattr(update_res, "data", None) or []
            if updated_rows:
                deducted = True
                print(
                    f"[Credits] Deducted {total} credits for job {job_id} "
                    f"(monthly: {monthly_deducted}, addon: {addon_deducted}, source: {deduction_source})"
                )
                break

            time.sleep(0.1)

        if not deducted:
            raise RuntimeError("Failed to atomically deduct credits")

        # Record in ledger with deduction source
        supabase_client.table("ledger").insert(
            {
                "user_id": user_id,
                "change": -total,
                "amount": 0.0,  # ensure non-null for deductions
                "reason": f"job deduction: {job_id} ({deduction_source})",
                "ts": datetime.utcnow().isoformat(),
            }
        ).execute()

        # Store breakdown in meta for potential refunds
        meta["credit_cost"] = total
        meta["monthly_deducted"] = monthly_deducted
        meta["addon_deducted"] = addon_deducted
        meta["credits_deducted"] = True
        if "credits_refunded" not in meta:
            meta["credits_refunded"] = False
        supabase_client.table("jobs").update({"meta_json": meta}).eq("id", job_id).execute()

        return True

    except Exception as exc:
        print(f"[Credits] Error while deducting for job {job_id}: {exc}")
        if deducted:
            # Rollback: restore both buckets
            try:
                supabase_client.table("profiles").update({
                    "credits_remaining": previous_monthly,
                    "addon_credits": previous_addon
                }).eq("id", user_id).execute()
                print(f"[Credits] Rolled back deduction for job {job_id}")
            except Exception as rollback_exc:
                print(
                    f"[Credits] Failed to rollback deduction for job {job_id}: {rollback_exc}"
                )
        supabase_client.table("jobs").update(
            {
                "status": "failed",
                "finished_at": datetime.utcnow().isoformat() + "Z",
                "error": "Unable to deduct credits",
            }
        ).eq("id", job_id).execute()
        return False
    finally:
        # Check for welcome reward unlock (500 credits used within 7 days)
        if deducted:
            try:
                # 1. Check current status
                p_res = supabase_client.table("profiles").select("created_at, welcome_reward_status").eq("id", user_id).single().execute()
                if p_res.data:
                    status = p_res.data.get("welcome_reward_status")
                    created_at_str = p_res.data.get("created_at")
                    
                    if status == "locked" and created_at_str:
                        # 2. Check time constraint (7 days)
                        created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                        if datetime.now(created_at.tzinfo) < created_at + timedelta(days=7):
                            # 3. Check total usage
                            l_res = supabase_client.rpc(
                                "get_total_credits_used", 
                                {"user_uuid": user_id}
                            ).execute()
                            
                            total_used = 0
                            if hasattr(l_res, 'data') and l_res.data is not None:
                                total_used = l_res.data
                            
                            if total_used >= 500:
                                supabase_client.table("profiles").update({
                                    "welcome_reward_status": "unlocked"
                                }).eq("id", user_id).execute()
                                print(f"[Rewards] Unlocked welcome reward for user {user_id}")

            except Exception as e:
                print(f"[Rewards] Failed to check reward status: {e}")


def _get_job_timeout():
    try:
        return int(os.getenv("SUBJOB_TIMEOUT", "600"))
    except (TypeError, ValueError):
        return 600


def _update_job_progress(
    job_id: str,
    total_rows: int,
    processed_in_chunk: int,
    last_reported: int,
    *,
    supabase_client=supabase,
    max_attempts: int = 5,
    retry_delay: float = 0.05,
) -> Tuple[int, Optional[dict]]:
    """Attempt to update global job progress using optimistic concurrency control."""

    delta = processed_in_chunk - last_reported
    if delta <= 0:
        return last_reported, None

    attempts = 0
    while attempts < max_attempts:
        job_res = (
            supabase_client.table("jobs")
            .select("rows_processed")
            .eq("id", job_id)
            .limit(1)
            .execute()
        )
        already_done = job_res.data[0]["rows_processed"] if job_res.data else 0
        new_done = min(total_rows, already_done + delta)
        percent = round((new_done / total_rows) * 100, 2) if total_rows else 0.0

        update_res = (
            supabase_client.table("jobs")
            .update({"rows_processed": new_done, "progress_percent": percent})
            .eq("id", job_id)
            .eq("rows_processed", already_done)
            .execute()
        )

        if getattr(update_res, "data", None):
            return processed_in_chunk, {
                "new_done": new_done,
                "percent": percent,
                "delta": delta,
            }

        attempts += 1
        print(
            f"[Worker] Job {job_id} progress update conflict (attempt {attempts}/{max_attempts}); retrying"
        )
        if retry_delay:
            time.sleep(retry_delay)

    raise RuntimeError(
        f"Failed to update progress for job {job_id} after {max_attempts} attempts"
    )

# -----------------------------
# Timing helper
# -----------------------------
def record_time(label, start, job_id):
    elapsed = round(time.time() - start, 2)
    print(f"[Timing] {label} took {elapsed} sec for job {job_id}")
    return elapsed

# -----------------------------
# Job Worker Functions
# -----------------------------

def next_queued_job():
    """Fetch the next queued job from Supabase."""
    res = (
        supabase.table("jobs")
        .select("*")
        .eq("status", "queued")
        .order("created_at", desc=False)
        .limit(1)
        .execute()
    )
    return res.data[0] if res.data else None


def _process_small_job_inline(
    job_id: str,
    user_id: str,
    local_path: str,
    headers: List[str],
    total: int,
    row_iter,
    meta: dict,
    final_output_headers: List[str],
    timings: dict,
    job_start: float,
):
    """
    Process small jobs (<100 rows) inline without chunking.

    This fast path eliminates 60+ storage operations by processing rows
    directly in the dispatcher worker instead of creating chunks and RQ subjobs.
    """
    inline_start = time.time()
    print(f"[Worker] Job {job_id} | Using SMALL-FILE FAST PATH for {total} rows (inline processing)")

    # Resolve headers for output
    row_headers, email_header, _, _ = _resolve_output_header_order(headers, meta)

    # Load all rows into memory
    rows = list(row_iter)
    print(f"[Worker] Job {job_id} | Processing {len(rows)} rows in parallel (inline, no chunks)")

    # Process all rows in parallel
    results = []
    with ThreadPoolExecutor(max_workers=PARALLEL_ROWS_PER_WORKER) as executor:
        futures = {}
        for i, row in enumerate(rows):
            future = executor.submit(
                _process_single_row,
                i,
                row,
                row_headers,
                email_header,
                meta,
                job_id,
                0,  # chunk_id (not used for inline)
            )
            futures[future] = i

        # Collect results as they complete
        for future in as_completed(futures):
            try:
                result = future.result()
                results.append(result)
            except Exception as exc:
                row_idx = futures[future]
                print(f"[Worker] Job {job_id} | Inline row {row_idx} failed: {exc}")
                # Add error row
                error_row = {header: "" for header in row_headers}
                if email_header:
                    error_row[email_header] = ""
                error_row["email_body"] = f"Error: {str(exc)}"
                error_row["sif_personalized_line"] = ""
                results.append((row_idx, error_row, str(exc)))

    # Sort by original row index
    results.sort(key=lambda x: x[0])

    timings["inline_processing"] = record_time(f"Inline processing ({total} rows)", inline_start, job_id)

    # Write final result
    output_start = time.time()
    final_df = pd.DataFrame([normalized_row for _, normalized_row, _ in results])

    # Ensure all output headers are present
    for header in final_output_headers:
        if header not in final_df.columns:
            final_df[header] = ""

    # Reorder columns to match expected output
    ordered_cols = [col for col in final_output_headers if col in final_df.columns]
    extra_cols = [col for col in final_df.columns if col not in ordered_cols]
    final_df = final_df[ordered_cols + extra_cols]

    # Write to temp files
    local_dir = tempfile.mkdtemp()
    out_xlsx = os.path.join(local_dir, f"{job_id}_final.xlsx")
    try:
        final_df.to_excel(out_xlsx, index=False, engine="openpyxl")

        # Upload final result
        storage_path = f"{user_id}/{job_id}/result.xlsx"
        with open(out_xlsx, "rb") as f:
            _upload_to_storage(storage_path, f, f"inline result for job {job_id}")

        timings["inline_output"] = record_time("Write and upload final result", output_start, job_id)
        timings["process_job_total"] = record_time("process_job total (inline)", job_start, job_id)

        # Mark job as succeeded
        supabase.table("jobs").update(
            {
                "status": "succeeded",
                "finished_at": datetime.utcnow().isoformat() + "Z",
                "result_path": storage_path,
                "progress_percent": 100,
                "rows_processed": total,
                "timing_json": json.dumps(timings),
            }
        ).eq("id", job_id).execute()

        print(f"[Worker] Job {job_id} | Completed inline processing successfully")

    finally:
        shutil.rmtree(local_dir, ignore_errors=True)


def _process_single_row(
    row_index: int,
    row: dict,
    row_headers: List[str],
    email_header: Optional[str],
    meta: dict,
    job_id: str,
    chunk_id: int,
) -> Tuple[int, dict, Optional[str]]:
    """
    Process a single row in a thread.

    Returns:
        tuple: (row_index, normalized_row_dict, error_message)
    """
    try:
        email_value = row.get(email_header, "") if email_header else ""

        # Perform research
        research_components = "Research unavailable: unexpected error."
        try:
            research_components = perform_research(email_value)
        except Exception as research_exc:
            error_msg = f"Research error: {research_exc}"
            print(f"[Worker] Job {job_id} | Chunk {chunk_id} | Row {row_index + 1} | {error_msg}")
            research_components = f"Research unavailable: {str(research_exc)}"

        # Generate email body
        email_body = "Email body unavailable: unexpected error."
        try:
            service_context = meta.get("service", "{}")
            email_body = generate_full_email_body(
                research_components,
                service_context,
            )
            # Apply cleaning pipeline
            email_body = clean_email_body(email_body)
        except Exception as email_exc:
            error_msg = f"Email generation error: {email_exc}"
            print(f"[Worker] Job {job_id} | Chunk {chunk_id} | Row {row_index + 1} | {error_msg}")
            email_body = f"Email unavailable: {str(email_exc)}"

        # Build normalized row
        normalized_row = {}
        for header in row_headers:
            value = row.get(header, "")
            normalized_row[header] = "" if value is None else value

        if email_header:
            normalized_row[email_header] = "" if email_value is None else email_value

        normalized_row["email_body"] = email_body

        # Extract first paragraph for sif_personalized_line
        paragraphs = email_body.split('\n\n')
        first_paragraph = paragraphs[0].strip() if paragraphs else ""
        normalized_row["sif_personalized_line"] = first_paragraph

        return (row_index, normalized_row, None)

    except Exception as exc:
        # Return error row if anything fails
        error_msg = f"Row processing error: {exc}"
        print(f"[Worker] Job {job_id} | Chunk {chunk_id} | Row {row_index + 1} | CRITICAL ERROR: {error_msg}")
        traceback.print_exc()

        # Build error row with empty values
        error_row = {}
        for header in row_headers:
            error_row[header] = row.get(header, "")
        if email_header:
            error_row[email_header] = row.get(email_header, "")
        error_row["email_body"] = f"Error: {str(exc)}"
        error_row["sif_personalized_line"] = ""

        return (row_index, error_row, str(exc))


def process_subjob(job_id: str, chunk_id: int, chunk_storage_path: str, meta: dict, user_id: str, total_rows: int):
    """Process a chunk of rows for a given job, with global progress logging."""
    sub_start = time.time()
    processed_in_chunk = 0
    rows_since_last_report = 0
    last_reported = 0
    chunk_input_path = _chunk_raw_local_path(job_id, chunk_id)
    downloaded_temp_dir = None
    cleanup_local_raw = False

    try:
        if os.path.exists(chunk_input_path):
            print(
                f"[Worker] Job {job_id} | Chunk {chunk_id} | using local raw chunk at {chunk_input_path}"
            )
        else:
            print(
                f"[Worker] Job {job_id} | Chunk {chunk_id} | downloading raw chunk {chunk_storage_path}"
            )
            downloaded_path = _download_chunk_from_storage(chunk_storage_path)
            downloaded_temp_dir = os.path.dirname(downloaded_path)
            chunk_input_path = downloaded_path

        headers, chunk_total_rows = _csv_headers_and_total(chunk_input_path)
        if chunk_total_rows == 0:
            print(f"[Worker] Chunk {chunk_id} for job {job_id} is empty; skipping generation")
            _remove_from_storage(
                chunk_storage_path,
                f"raw chunk {chunk_id} for job {job_id}",
                bucket=RAW_CHUNK_BUCKET,
            )
            cleanup_local_raw = True
            return None

        local_dir = os.path.join("/data/chunks", job_id)
        os.makedirs(local_dir, exist_ok=True)
        out_path = os.path.join(local_dir, f"chunk_{chunk_id}.csv")

        # Read input CSV and prepare for parallel processing
        with open(chunk_input_path, newline="", encoding="utf-8-sig") as in_f:
            reader = csv.DictReader(in_f)
            input_headers = reader.fieldnames or headers
            meta = _ensure_dict(meta)
            row_headers, email_header, output_headers, _ = _resolve_output_header_order(
                input_headers, meta
            )
            # Load all rows into memory for parallel processing
            rows = list(reader)

        print(f"[Worker] Job {job_id} | Chunk {chunk_id} | Processing {len(rows)} rows in parallel with {PARALLEL_ROWS_PER_WORKER} workers")

        # Parallel processing with ThreadPoolExecutor
        results = []
        completed_count = 0
        rows_since_last_report = 0
        last_reported = 0
        progress_lock = Lock()

        with ThreadPoolExecutor(max_workers=PARALLEL_ROWS_PER_WORKER) as executor:
            # Submit all rows to thread pool
            futures = {}
            for i, row in enumerate(rows):
                future = executor.submit(
                    _process_single_row,
                    i,  # row_index
                    row,  # row dict
                    row_headers,
                    email_header,
                    meta,
                    job_id,
                    chunk_id,
                )
                futures[future] = i

            # Collect results as they complete
            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as exc:
                    # This should never happen because _process_single_row catches everything
                    row_idx = futures[future]
                    print(f"[Worker] Job {job_id} | Chunk {chunk_id} | CRITICAL: Thread {row_idx} exception: {exc}")
                    traceback.print_exc()
                    # Add error result
                    error_row = {header: "" for header in row_headers}
                    if email_header:
                        error_row[email_header] = ""
                    error_row["email_body"] = f"Critical error: {str(exc)}"
                    error_row["sif_personalized_line"] = ""
                    results.append((row_idx, error_row, str(exc)))

                # Update progress atomically
                with progress_lock:
                    completed_count += 1
                    rows_since_last_report += 1
                    current = completed_count
                    rows_since = rows_since_last_report

                # Update progress every 5 rows or on last row
                is_last_row = current == chunk_total_rows
                if rows_since >= 5 or is_last_row:
                    try:
                        last_reported, progress_info = _update_job_progress(
                            job_id,
                            total_rows,
                            current,
                            last_reported,
                        )
                    except RuntimeError as exc:
                        print(
                            f"[Worker] Job {job_id} | Chunk {chunk_id} | Progress update failed: {exc}"
                        )
                    else:
                        with progress_lock:
                            rows_since_last_report = 0
                        if progress_info:
                            new_done = progress_info["new_done"]
                            percent = progress_info["percent"]
                            delta = progress_info["delta"]
                            print(
                                f"[Worker] Job {job_id} | Chunk {chunk_id} | Progress +{delta} -> {new_done}/{total_rows} rows ({percent}%)"
                            )
                            # Insert progress into job_logs (for history/recovery)
                            supabase.table("job_logs").insert(
                                {
                                    "job_id": job_id,
                                    "step": new_done,
                                    "total": total_rows,
                                    "message": (
                                        f"Global progress: +{delta} rows -> {new_done}/{total_rows} ({percent}%)"
                                    ),
                                }
                            ).execute()

                            # Publish real-time progress update to Redis pub/sub for WebSocket
                            try:
                                # Get current job status from database
                                job_status_res = supabase.table("jobs").select("status").eq("id", job_id).single().execute()
                                current_status = job_status_res.data.get("status", "in_progress") if job_status_res.data else "in_progress"

                                progress_data = {
                                    "job_id": job_id,
                                    "status": current_status,
                                    "percent": percent,
                                    "message": f"Global progress: +{delta} rows -> {new_done}/{total_rows} ({percent}%)",
                                }
                                redis_conn.publish(f"job_progress:{job_id}", json.dumps(progress_data))
                                print(f"[Worker] Published progress for job {job_id}: {percent}% to Redis channel job_progress:{job_id}")
                            except Exception as pub_error:
                                # Non-critical: WebSocket clients will fall back to polling
                                print(f"[Worker] Job {job_id} | Failed to publish progress to Redis: {pub_error}")

        # Sort results by original row index to preserve order
        results.sort(key=lambda x: x[0])

        # Write all results to output CSV (sequential, thread-safe)
        with open(out_path, "w", newline="", encoding="utf-8") as out_f:
            writer = csv.DictWriter(out_f, fieldnames=output_headers)
            writer.writeheader()

            for row_index, normalized_row, error in results:
                writer.writerow(normalized_row)
                if error:
                    print(f"[Worker] Job {job_id} | Chunk {chunk_id} | Row {row_index + 1} had error: {error}")

        print(f"[Worker] Saved local chunk {chunk_id} at {out_path}")

        storage_path = f"{user_id}/{job_id}/chunk_{chunk_id}.csv"
        with open(out_path, "rb") as f:
            _upload_to_storage(storage_path, f, f"chunk {chunk_id} for job {job_id}")

        supabase.table("files").insert(
            {
                "user_id": user_id,
                "job_id": job_id,
                "original_name": f"chunk_{chunk_id}.csv",
                "storage_path": storage_path,
                "file_type": "partial_output",
            }
        ).execute()

        try:
            rq_job = get_current_job()
            if rq_job:
                rq_job.meta[f"local_chunk_{chunk_id}"] = out_path
                rq_job.meta[f"storage_chunk_{chunk_id}"] = storage_path
                rq_job.save_meta()
                print(f"[Worker] Saved local path for chunk {chunk_id} -> {out_path}")
        except Exception as e:
            print(f"[Worker] Could not save local path for chunk {chunk_id}: {e}")

        elapsed = record_time(f"Chunk {chunk_id} row processing", sub_start, job_id)

        job_record = (
            supabase.table("jobs").select("timing_json").eq("id", job_id).limit(1).execute()
        )
        timings = {}
        if job_record.data and job_record.data[0].get("timing_json"):
            try:
                timings = json.loads(job_record.data[0]["timing_json"])
            except Exception:
                timings = {}
        if "chunks" not in timings:
            timings["chunks"] = {}
        timings["chunks"][str(chunk_id)] = elapsed
        supabase.table("jobs").update({"timing_json": json.dumps(timings)}).eq("id", job_id).execute()

        print(f"[Worker] Finished chunk {chunk_id}/{chunk_total_rows} for job {job_id}")

        _remove_from_storage(chunk_storage_path, f"raw chunk {chunk_id} for job {job_id}", bucket=RAW_CHUNK_BUCKET)
        cleanup_local_raw = True

        return storage_path

    except Exception as exc:
        error_message = f"Chunk {chunk_id} failed: {exc}"
        print(f"[Worker] Chunk error for job {job_id}: {exc}")
        traceback.print_exc()
        try:
            supabase.table("job_logs").insert(
                {
                    "job_id": job_id,
                    "step": chunk_id,
                    "total": total_rows,
                    "message": error_message,
                }
            ).execute()
        except Exception as log_exc:
            print(f"[Worker] Failed to log chunk error for job {job_id}: {log_exc}")

        try:
            supabase.table("jobs").update(
                {
                    "status": "failed",
                    "error": error_message,
                    "finished_at": datetime.utcnow().isoformat() + "Z",
                }
            ).eq("id", job_id).execute()
        except Exception as update_exc:
            print(f"[Worker] Failed to mark job {job_id} as failed: {update_exc}")

        refund_job_credits(job_id, user_id, "chunk error")
        raise
    finally:
        if cleanup_local_raw:
            try:
                if downloaded_temp_dir:
                    shutil.rmtree(downloaded_temp_dir, ignore_errors=True)
                elif os.path.exists(chunk_input_path):
                    os.remove(chunk_input_path)
                    parent_dir = os.path.dirname(chunk_input_path)
                    try:
                        if os.path.isdir(parent_dir) and not os.listdir(parent_dir):
                            os.rmdir(parent_dir)
                    except Exception:
                        pass
            except Exception as exc:
                print(
                    f"[Worker] Warning: failed to remove local raw chunk {chunk_id} for job {job_id}: {exc}"
                )


def finalize_job(
    job_id: str,
    user_id: str,
    total_chunks: int,
    final_headers: Optional[List[str]] = None,
):
    """Merge all partial CSV files into one final result and update job status."""
    finalize_start = time.time()
    try:
        print(f"[Worker] Finalizing job {job_id} with {total_chunks} chunks")

        # Load existing timing_json from DB
        job_record = supabase.table("jobs").select("timing_json").eq("id", job_id).limit(1).execute()
        timings = {}
        if job_record.data and job_record.data[0].get("timing_json"):
            try:
                timings = json.loads(job_record.data[0]["timing_json"])
            except Exception:
                timings = {}
        if "chunks" not in timings:
            timings["chunks"] = {}

        # --- Merge CSVs ---
        merge_start = time.time()
        frames = []
        for chunk_id in range(1, total_chunks + 1):
            local_path = os.path.join("/data/chunks", job_id, f"chunk_{chunk_id}.csv")
            if os.path.exists(local_path):
                print(f"[Worker] Using local chunk {chunk_id} for job {job_id}")
                df = pd.read_csv(local_path)
            else:
                print(f"[Worker] Local chunk {chunk_id} missing, downloading from Supabase...")
                storage_path = f"{user_id}/{job_id}/chunk_{chunk_id}.csv"
                signed = supabase.storage.from_("outputs").create_signed_url(storage_path, 300)
                url = signed.get("signedURL") if signed else None
                if not url:
                    raise Exception(f"Missing signed URL for {storage_path}")
                resp = requests.get(url, timeout=60)
                resp.raise_for_status()
                tmp_dir = tempfile.mkdtemp()
                local_path = os.path.join(tmp_dir, f"chunk_{chunk_id}.csv")
                with open(local_path, "wb") as f:
                    f.write(resp.content)
                df = pd.read_csv(local_path)

            frames.append(df)


        timings["merge_csvs"] = record_time("Merging CSV chunks", merge_start, job_id)

        # --- Final CSV → XLSX ---
        upload_start = time.time()
        final_df = pd.concat(frames, ignore_index=True)
        if final_headers:
            ordered_headers: List[str] = []
            seen_headers = set()
            for header in final_headers:
                if not header or header in seen_headers:
                    continue
                ordered_headers.append(header)
                seen_headers.add(header)
            for header in ordered_headers:
                if header not in final_df.columns:
                    final_df[header] = ""
            ordered_in_df = [header for header in ordered_headers if header in final_df.columns]
            extra_headers = [header for header in final_df.columns if header not in ordered_in_df]
            final_df = final_df[ordered_in_df + extra_headers]
        local_dir = tempfile.mkdtemp()
        out_csv = os.path.join(local_dir, f"{job_id}_final.csv")
        out_xlsx = os.path.join(local_dir, f"{job_id}_final.xlsx")

        final_df.to_csv(out_csv, index=False)
        final_df.to_excel(out_xlsx, index=False, engine="openpyxl")

        storage_path = f"{user_id}/{job_id}/result.xlsx"
        with open(out_xlsx, "rb") as f:
            _upload_to_storage(storage_path, f, f"final result for job {job_id}")
        timings["csv_to_xlsx"] = record_time("Final CSV→XLSX + upload", upload_start, job_id)

        timings["finalize_total"] = record_time("Finalize total", finalize_start, job_id)

        # Save full timings
        supabase.table("jobs").update(
            {
                "status": "succeeded",
                "finished_at": datetime.utcnow().isoformat() + "Z",
                "result_path": storage_path,
                "progress_percent": 100,
                "timing_json": json.dumps(timings),
            }
        ).eq("id", job_id).execute()

        # Publish final success status to Redis pub/sub for WebSocket
        try:
            success_data = {
                "job_id": job_id,
                "status": "succeeded",
                "percent": 100,
                "message": "Job completed successfully",
            }
            redis_conn.publish(f"job_progress:{job_id}", json.dumps(success_data))
            print(f"[Worker] Published success status for job {job_id}")
        except Exception as pub_error:
            print(f"[Worker] Failed to publish success status to Redis: {pub_error}")

        # Cleanup local chunks
        local_job_dir = os.path.join("/data/chunks", job_id)
        if os.path.exists(local_job_dir):
            import shutil
            shutil.rmtree(local_job_dir, ignore_errors=True)
            print(f"[Worker] Cleaned up local chunks for job {job_id}")


        print(f"[Worker] Finalized job {job_id} successfully")

    except Exception as e:
        print(f"[Worker] Finalize ERROR for job {job_id}: {e}")
        traceback.print_exc()
        error_message = f"Finalize error: {e}"
        supabase.table("jobs").update(
            {"status": "failed", "error": error_message}
        ).eq("id", job_id).execute()

        # Publish failure status to Redis pub/sub for WebSocket
        try:
            failure_data = {
                "job_id": job_id,
                "status": "failed",
                "percent": 0,
                "message": error_message,
            }
            redis_conn.publish(f"job_progress:{job_id}", json.dumps(failure_data))
            print(f"[Worker] Published failure status for job {job_id}")
        except Exception as pub_error:
            print(f"[Worker] Failed to publish failure status to Redis: {pub_error}")

        refund_job_credits(job_id, user_id, "finalize error")


def process_job(job_id: str):
    """Split a job into chunks based on available workers and enqueue subjobs."""
    job_start = time.time()
    timings = {}
    job = None
    try:
        print(f"[Worker] === Starting process_job for {job_id} ===")

        job_res = (
            supabase.table("jobs")
            .select("*")
            .eq("id", job_id)
            .limit(1)
            .execute()
        )
        if not job_res.data:
            print(f"[Worker] Job {job_id} not found in DB")
            return
        job = job_res.data[0]

        meta = _ensure_dict(job.get("meta_json"))

        user_id = job["user_id"]
        file_path = meta.get("file_path")
        if not file_path:
            supabase.table("jobs").update(
                {"status": "failed", "error": "Missing file_path"}
            ).eq("id", job_id).execute()
            return

        # --- Download file ---
        dl_start = time.time()
        signed = supabase.storage.from_("inputs").create_signed_url(file_path, 300)
        url = signed.get("signedURL") if signed else None
        if not url:
            supabase.table("jobs").update(
                {
                    "status": "failed",
                    "finished_at": datetime.utcnow().isoformat() + "Z",
                    "error": "Could not create signed URL",
                }
            ).eq("id", job_id).execute()
            return

        local_dir = tempfile.mkdtemp()
        local_path = os.path.join(local_dir, os.path.basename(file_path))
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        with open(local_path, "wb") as f:
            f.write(resp.content)

        # Optimize: Use cached row count from metadata if available (saves 2-10 seconds)
        cached_total = meta.get("total_rows")
        if cached_total is not None:
            print(f"[Worker] Job {job_id} | Using cached row count: {cached_total} (skipping re-count)")
            headers, total, row_iter = _get_headers_and_iterator(local_path, cached_total)
        else:
            # Legacy path: count rows manually (backwards compatibility)
            print(f"[Worker] Job {job_id} | No cached row count found; counting rows from file")
            headers, total, row_iter = _input_iterator(local_path)

        _, _, final_output_headers, chunk_headers = _resolve_output_header_order(
            headers, meta
        )

        timings["download_input"] = record_time("Download input file", dl_start, job_id)

        # --- Claim job (atomic via database) ---
        setup_start = time.time()

        # Verify job is still queued and get latest metadata
        refreshed_job = (
            supabase.table("jobs")
            .select("status, meta_json")
            .eq("id", job_id)
            .limit(1)
            .execute()
        )
        if not refreshed_job.data:
            print(f"[Worker] Job {job_id} disappeared before claiming; aborting")
            return

        current_row = refreshed_job.data[0]
        current_status = current_row.get("status")
        meta = _ensure_dict(current_row.get("meta_json"))
        job["status"] = current_status
        job["meta_json"] = meta

        if current_status != "queued":
            print(f"[Worker] Job {job_id} is already {current_status}; skipping claim")
            return

        # Optimize: Skip redundant credit deduction if already done in API
        if meta.get("credits_deducted"):
            print(f"[Worker] Job {job_id} credits already deducted in API; skipping worker deduction")
            should_continue = True
        else:
            # Legacy path: deduct credits if not already done (backwards compatibility)
            # Note: _deduct_job_credits has its own internal locking for credit safety
            print(f"[Worker] Job {job_id} credits not yet deducted; processing deduction")
            should_continue = _deduct_job_credits(
                job_id, user_id, total, meta, supabase_client=supabase
            )

        if not should_continue:
            print(f"[Worker] Skipping job {job_id} after credit check")
            return

        # Atomic claim: Only succeeds if job is still queued
        # If another worker already claimed it, this update returns 0 rows
        claim_payload = {
            "status": "in_progress",
            "started_at": datetime.utcnow().isoformat() + "Z",
            "rows_processed": 0,
            "progress_percent": 0,
        }
        claim_res = (
            supabase.table("jobs")
            .update(claim_payload)
            .eq("id", job_id)
            .eq("status", "queued")  # Atomic: only update if still queued
            .execute()
        )
        updated_rows = getattr(claim_res, "data", None) or []
        if not updated_rows:
            print(f"[Worker] Job {job_id} was claimed by another worker; aborting")
            # Only refund if we deducted credits (legacy path)
            if not meta.get("credits_deducted"):
                refund_job_credits(job_id, user_id, "claim failed")
            return

        job.update(updated_rows[0])
        job["status"] = claim_payload["status"]
        job["started_at"] = claim_payload["started_at"]
        job["rows_processed"] = claim_payload["rows_processed"]
        job["progress_percent"] = claim_payload["progress_percent"]
        job["meta_json"] = meta

        timings["setup"] = record_time("Setup (DB updates + job claim)", setup_start, job_id)

        # --- Small-file fast path: Process inline for files < 100 rows ---
        SMALL_FILE_THRESHOLD = 100
        if total > 0 and total < SMALL_FILE_THRESHOLD:
            print(f"[Worker] Job {job_id} | Small file detected ({total} rows < {SMALL_FILE_THRESHOLD}), using inline processing")
            _process_small_job_inline(
                job_id,
                user_id,
                local_path,
                headers,
                total,
                row_iter,
                meta,
                final_output_headers,
                timings,
                job_start,
            )
            return  # Job complete, skip chunking

        # --- Chunking (Parallelized) ---
        chunk_start = time.time()
        try:
            num_workers = int(os.getenv("WORKER_COUNT", "1"))
        except Exception:
            num_workers = 1

        subjob_refs = []
        chunk_buffer = []
        chunks_to_persist = []  # List of (chunk_id, rows) tuples
        chunk_count = 0

        job_timeout = _get_job_timeout()

        if total > 0:
            num_chunks = min(total, num_workers) or 1
            chunk_size = max(1, math.ceil(total / num_chunks))

            # Phase 1: Collect all chunks in memory
            print(f"[Worker] Job {job_id} | Collecting chunks for parallel persistence...")
            for row in row_iter:
                chunk_buffer.append(row)
                if len(chunk_buffer) >= chunk_size:
                    chunk_count += 1
                    # Store chunk data for later parallel persistence
                    chunks_to_persist.append((chunk_count, list(chunk_buffer)))
                    chunk_buffer = []

            # Handle remaining rows
            if chunk_buffer:
                chunk_count += 1
                chunks_to_persist.append((chunk_count, list(chunk_buffer)))
                chunk_buffer = []

            # Phase 2: Persist all chunks in parallel
            def persist_chunk_task(chunk_id, rows):
                """Persist a single chunk to storage (runs in thread pool)."""
                storage_path = _persist_chunk_rows(
                    job_id,
                    chunk_id,
                    chunk_headers,
                    rows,
                    user_id,
                )
                return (chunk_id, storage_path)

            chunk_storage_paths = {}  # Map chunk_id -> storage_path

            if chunks_to_persist:
                # Use ThreadPoolExecutor to persist chunks in parallel
                max_parallel_uploads = min(len(chunks_to_persist), 10)  # Cap at 10 concurrent uploads
                print(f"[Worker] Job {job_id} | Persisting {len(chunks_to_persist)} chunks in parallel (max {max_parallel_uploads} workers)...")

                persist_start = time.time()
                with ThreadPoolExecutor(max_workers=max_parallel_uploads) as executor:
                    futures = {
                        executor.submit(persist_chunk_task, chunk_id, rows): chunk_id
                        for chunk_id, rows in chunks_to_persist
                    }

                    for future in as_completed(futures):
                        chunk_id, storage_path = future.result()
                        chunk_storage_paths[chunk_id] = storage_path

                timings["parallel_chunk_persistence"] = record_time(
                    f"Parallel chunk persistence ({len(chunks_to_persist)} chunks)",
                    persist_start,
                    job_id
                )

                # Phase 3: Enqueue all subjobs (in order)
                print(f"[Worker] Job {job_id} | Enqueuing {len(chunk_storage_paths)} subjobs...")
                for chunk_id in sorted(chunk_storage_paths.keys()):
                    storage_path = chunk_storage_paths[chunk_id]
                    job_ref = queue.enqueue(
                        process_subjob,
                        job_id,
                        chunk_id,
                        storage_path,
                        meta,
                        user_id,
                        total,
                        job_timeout=job_timeout,
                    )
                    subjob_refs.append(job_ref)

        timings["chunking_total"] = record_time("Chunking + parallel persist + enqueue", chunk_start, job_id)

        if chunk_count > 0:
            queue.enqueue(
                finalize_job,
                job_id,
                user_id,
                chunk_count,
                final_output_headers,
                depends_on=subjob_refs,
                job_timeout=job_timeout,
            )
        else:
            _finalize_empty_job(
                job_id,
                user_id,
                final_output_headers,
                timings,
                job_start,
            )
            return

        timings["process_job_total"] = record_time("process_job total", job_start, job_id)
        supabase.table("jobs").update({"timing_json": json.dumps(timings)}).eq("id", job_id).execute()

        print(f"[Worker] Enqueued {chunk_count} chunks + finalize for job {job_id}")

    except Exception as e:
        print(f"[Worker] FATAL ERROR job {job_id}: {e}")
        traceback.print_exc()
        supabase.table("jobs").update(
            {
                "status": "failed",
                "finished_at": datetime.utcnow().isoformat() + "Z",
                "error": str(e),
            }
        ).eq("id", job_id).execute()
        user_for_refund = job.get("user_id") if isinstance(job, dict) else None
        refund_job_credits(job_id, user_for_refund, "process_job error")


def worker_loop(poll_interval: int = 30):
    """
    Worker loop that uses Redis pub/sub for instant job notifications with fallback polling.

    The pub/sub approach eliminates the 0-5 second delay from polling, allowing jobs to start
    immediately when enqueued. Fallback polling (every 30s) ensures reliability.
    """
    print("[Worker] Starting event-driven dispatcher loop with Redis pub/sub...")

    # Create a separate Redis connection for pub/sub (blocking operations)
    pubsub_conn = redis.Redis(host="redis", port=6379, decode_responses=True)
    pubsub = pubsub_conn.pubsub(ignore_subscribe_messages=True)

    # Subscribe to job notification channel
    pubsub.subscribe(JOB_NOTIFICATION_CHANNEL)

    print(f"[Worker] Subscribed to Redis channel: {JOB_NOTIFICATION_CHANNEL}")
    print(f"[Worker] Fallback polling interval: {poll_interval}s")

    while True:
        try:
            # Listen for pub/sub messages with timeout (non-blocking with fallback)
            # This allows us to check for jobs every poll_interval seconds even without notifications
            message = pubsub.get_message(timeout=poll_interval)

            if message and message['type'] == 'message':
                job_id = message['data']
                print(f"[Worker] Received job notification for job_id: {job_id}")

            # Always check for queued jobs (handles both notified and missed jobs)
            job = next_queued_job()
            if not job:
                if not message:  # Only log if we polled (not notified)
                    print("[Worker] No jobs found (polling)...")
                continue

            print(f"[Worker] Processing job {job['id']}")
            process_job(job["id"])

        except redis.exceptions.ConnectionError as e:
            print(f"[Worker] Redis connection error: {e}. Falling back to polling...")
            time.sleep(poll_interval)
            # Try to reconnect pub/sub
            try:
                pubsub.close()
                pubsub_conn = redis.Redis(host="redis", port=6379, decode_responses=True)
                pubsub = pubsub_conn.pubsub(ignore_subscribe_messages=True)
                pubsub.subscribe(JOB_NOTIFICATION_CHANNEL)
                print(f"[Worker] Reconnected to Redis channel: {JOB_NOTIFICATION_CHANNEL}")
            except Exception as reconnect_exc:
                print(f"[Worker] Failed to reconnect to pub/sub: {reconnect_exc}")
        except Exception as e:
            print("[Worker] Dispatcher ERROR:", str(e))
            traceback.print_exc()
            time.sleep(poll_interval)
