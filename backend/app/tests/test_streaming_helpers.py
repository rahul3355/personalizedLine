import csv
import os
from pathlib import Path

import asyncio

import pytest
from io import BytesIO
from openpyxl import Workbook, load_workbook

import sys
sys.path.append(str(Path(__file__).resolve().parents[3]))

from backend.app import file_streaming as file_streaming_module

from backend.app.file_streaming import (
    count_csv_rows,
    count_xlsx_rows,
    extract_csv_headers,
    extract_xlsx_headers,
)


@pytest.fixture()
def tmp_csv(tmp_path: Path) -> Path:
    path = tmp_path / "data.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["col_a", "col_b"])
        for i in range(5):
            writer.writerow([i, i + 1])
    return path


@pytest.fixture()
def tmp_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["col_a", "col_b"])
    for i in range(7):
        ws.append([i, i + 2])
    wb.save(path)
    wb.close()
    return path


def test_csv_helpers(tmp_csv: Path):
    assert extract_csv_headers(str(tmp_csv)) == ["col_a", "col_b"]
    assert count_csv_rows(str(tmp_csv)) == 5


def test_xlsx_helpers(tmp_xlsx: Path):
    assert extract_xlsx_headers(str(tmp_xlsx)) == ["col_a", "col_b"]
    assert count_xlsx_rows(str(tmp_xlsx)) == 7


def test_stream_input_to_tempfile_preserves_suffix(monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws.append(["col_a", "col_b"])
    for i in range(3):
        ws.append([i, i + 5])
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    payload = buffer.getvalue()

    class DummyBucket:
        def __init__(self):
            self.calls = []

        def create_signed_url(self, file_path, expires_in):
            self.calls.append((file_path, expires_in))
            return {"signedURL": "https://example.com/signed"}

    class DummyStorage:
        def __init__(self):
            self.bucket = DummyBucket()

        def from_(self, name):
            assert name == "inputs"
            return self.bucket

    class DummySupabase:
        def __init__(self):
            self.storage = DummyStorage()

    class DummyStream:
        status_code = 200

        def __init__(self, data):
            self._data = data

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def aiter_bytes(self, chunk_size):
            yield self._data

    class DummyAsyncClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        def stream(self, method, url):
            assert method == "GET"
            assert url == "https://example.com/signed"
            return DummyStream(payload)

    monkeypatch.setattr(
        file_streaming_module.httpx, "AsyncClient", DummyAsyncClient
    )

    monkeypatch.setenv("SUPABASE_URL", "https://project.supabase.co")

    supabase = DummySupabase()
    temp_path = asyncio.run(
        file_streaming_module.stream_input_to_tempfile(
            supabase, "user-id/data.xlsx"
        )
    )

    try:
        assert temp_path.endswith(".xlsx")
        workbook = load_workbook(temp_path, read_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            assert rows[0] == ("col_a", "col_b")
            assert rows[1] == (0, 5)
        finally:
            workbook.close()
    finally:
        os.unlink(temp_path)


def test_stream_input_to_tempfile_handles_relative_signed_url(monkeypatch):
    payload = b"test"

    class DummyBucket:
        def create_signed_url(self, file_path, expires_in):
            return {"signedURL": "/storage/v1/object/sign/inputs/data.csv"}

    class DummyStorage:
        def from_(self, name):
            return DummyBucket()

    class DummySupabase:
        def __init__(self):
            self.storage = DummyStorage()

    class DummyStream:
        status_code = 200

        def __init__(self, data):
            self._data = data

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def aiter_bytes(self, chunk_size):
            yield self._data

    class DummyAsyncClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        def stream(self, method, url):
            assert url == "https://project.supabase.co/storage/v1/object/sign/inputs/data.csv"
            return DummyStream(payload)

    monkeypatch.setattr(
        file_streaming_module.httpx, "AsyncClient", DummyAsyncClient
    )
    monkeypatch.setenv("SUPABASE_URL", "https://project.supabase.co")

    supabase = DummySupabase()
    temp_path = asyncio.run(
        file_streaming_module.stream_input_to_tempfile(supabase, "user/data.csv")
    )

    try:
        with open(temp_path, "rb") as handle:
            assert handle.read() == payload
    finally:
        os.unlink(temp_path)


def test_stream_input_to_tempfile_requires_supabase_url(monkeypatch):
    class DummyBucket:
        def create_signed_url(self, file_path, expires_in):
            return {"signedURL": "/relative"}

    class DummyStorage:
        def from_(self, name):
            return DummyBucket()

    class DummySupabase:
        def __init__(self):
            self.storage = DummyStorage()

    class DummyAsyncClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        def stream(self, method, url):
            raise AssertionError("Should not be called")

    monkeypatch.setattr(
        file_streaming_module.httpx, "AsyncClient", DummyAsyncClient
    )
    monkeypatch.delenv("SUPABASE_URL", raising=False)

    supabase = DummySupabase()

    with pytest.raises(file_streaming_module.FileStreamingError) as exc:
        asyncio.run(
            file_streaming_module.stream_input_to_tempfile(
                supabase, "user/data.csv"
            )
        )

    assert "SUPABASE_URL" in str(exc.value)
