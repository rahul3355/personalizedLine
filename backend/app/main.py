from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, Depends, Header, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from typing import Union, Optional, Dict
import uuid, shutil, os, tempfile, threading, json, stripe, time, re, asyncio
from fastapi.responses import StreamingResponse
import io
from pydantic import BaseModel
import os
import logging
from . import jobs
from .file_streaming import (
    FileStreamingError,
    count_csv_rows,
    count_xlsx_rows,
    extract_csv_headers,
    extract_xlsx_headers,
    stream_input_to_tempfile,
)
from .supabase_client import supabase
from supabase import create_client
from fastapi import APIRouter, Body
from io import BytesIO
from datetime import datetime, timedelta
import requests
from fastapi import Query
import redis
from rq import Queue
 # reuse the same function your workers use


# Load env
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))
stripe.api_key = os.getenv("STRIPE_SECRET_KEY")

# Change this value to trigger a deployment via GitHub Actions
DEPLOY_TRIGGER_V1 = "trigger_v4"

app = FastAPI()

import os
redis_url = os.getenv("REDIS_URL", "redis://redis:6379")
redis_conn = redis.from_url(redis_url, decode_responses=True)
  # use localhost when FastAPI runs on your host
q = Queue(connection=redis_conn)


# Security scheme (adds Authorize button in Swagger)
security = HTTPBearer()

from dataclasses import dataclass
from typing import Any, Dict, Optional

try:
    import jwt
    from jwt import (
        ExpiredSignatureError,
        InvalidAudienceError,
        InvalidIssuerError,
        InvalidTokenError,
    )
except ModuleNotFoundError:  # pragma: no cover - fallback for offline environments
    from backend.app import jwt_fallback as jwt
    from backend.app.jwt_fallback import (
        ExpiredSignatureError,
        InvalidAudienceError,
        InvalidIssuerError,
        InvalidTokenError,
    )


@dataclass
class AuthenticatedUser:
    user_id: str
    claims: Dict[str, Any]

class CheckoutRequest(BaseModel):
    plan: str
    addon: bool = False
    quantity: int = 1

# Stripe Setup
STRIPE_SECRET = os.getenv("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET", "")
if STRIPE_SECRET:
    stripe.api_key = STRIPE_SECRET

# Base plan price mapping (monthly subscriptions)
PLAN_PRICE_MAP = {
    "starter": os.getenv("STRIPE_PRICE_STARTER"),
    "growth": os.getenv("STRIPE_PRICE_GROWTH"),
    "pro": os.getenv("STRIPE_PRICE_PRO"),
}

# Annual plan price mapping (yearly subscriptions)
PLAN_PRICE_MAP_ANNUAL = {
    "starter_annual": os.getenv("STRIPE_PRICE_STARTER_ANNUAL"),
    "growth_annual": os.getenv("STRIPE_PRICE_GROWTH_ANNUAL"),
    "pro_annual": os.getenv("STRIPE_PRICE_PRO_ANNUAL"),
}

# Add-on product mapping
ADDON_PRICE_MAP = {
    "free": os.getenv("STRIPE_PRICE_ADDON_FREE"),
    "starter": os.getenv("STRIPE_PRICE_ADDON_STARTER"),
    "growth": os.getenv("STRIPE_PRICE_ADDON_GROWTH"),
    "pro": os.getenv("STRIPE_PRICE_ADDON_PRO"),
}

# Map Stripe price IDs back to plan names
# Annual plans map to base plan name (e.g., "starter_annual" -> "starter")
PRICE_TO_PLAN = {
    **{pid: plan for plan, pid in PLAN_PRICE_MAP.items() if pid},
    **{pid: plan.replace("_annual", "") for plan, pid in PLAN_PRICE_MAP_ANNUAL.items() if pid}
}

CREDITS_MAP = {
    "free": 500,
    "starter": 2000,
    "growth": 10000,
    "pro": 40000,
}

# Annual plans receive 12x monthly credits upfront (no monthly resets)
ANNUAL_CREDITS_MAP = {
    "free": 500 * 12,
    "starter": 2000 * 12,   # 24,000 credits upfront
    "growth": 10000 * 12,   # 120,000 credits upfront
    "pro": 40000 * 12,      # 480,000 credits upfront
}

# Rollover configuration per plan
ROLLOVER_RULES = {
    "free": {"enabled": False, "max_multiplier": 1},
    "starter": {"enabled": False, "max_multiplier": 1},
    "growth": {"enabled": True, "max_multiplier": 2},
    "pro": {"enabled": True, "max_multiplier": 2},
}

# Monthly plan prices in USD (for bonus credit calculation on upgrade)
PLAN_PRICES = {
    "free": 0,
    "starter": 49,
    "growth": 149,
    "pro": 499,
}

# Addon credit expiration (in months)
ADDON_EXPIRATION_MONTHS = {
    "free": 12,
    "starter": 6,
    "growth": 12,
    "pro": 12,
}

FRONTEND_BASE_URLS = os.getenv(
    "FRONTEND_BASE_URLS",
    "http://localhost:3000,http://127.0.0.1:3000",
)

DEFAULT_ALLOWED_ORIGINS = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
]

ALLOWED_ORIGINS = [url.strip().rstrip("/") for url in FRONTEND_BASE_URLS.split(",") if url.strip()]

if not ALLOWED_ORIGINS:
    env_value = os.getenv("ENV", "").lower()
    debug_value = os.getenv("DEBUG", "").lower()
    if env_value in {"prod", "production"} or debug_value in {"0", "false", "no", "off"}:
        logging.warning(
            "FRONTEND_BASE_URLS resolved to an empty list in a production-like environment."
        )

    # Always fall back to localhost origins for development if no explicit list was provided.
    ALLOWED_ORIGINS = DEFAULT_ALLOWED_ORIGINS

APP_BASE_URL = os.getenv("APP_BASE_URL", "https://senditfast.ai").rstrip("/")

# Validate APP_BASE_URL in production to prevent localhost redirect issues
env_value = os.getenv("ENV", "").lower()
is_production = env_value in {"prod", "production"}
if is_production and "localhost" in APP_BASE_URL.lower():
    logging.error(
        f"CRITICAL: APP_BASE_URL is set to '{APP_BASE_URL}' in production environment. "
        "This will cause payment redirects to fail. Please set APP_BASE_URL to your production domain."
    )
    # In production, this is a critical error - log it prominently
    print("=" * 80)
    print("CRITICAL CONFIGURATION ERROR")
    print(f"APP_BASE_URL contains 'localhost': {APP_BASE_URL}")
    print("Payment redirects will FAIL in production!")
    print("Please set APP_BASE_URL environment variable to your production domain.")
    print("=" * 80)

SUCCESS_RETURN_PATH = os.getenv("STRIPE_SUCCESS_PATH", "/billing/success")
SUCCESS_URL = f"{APP_BASE_URL}{SUCCESS_RETURN_PATH}"
CANCEL_URL = f"{APP_BASE_URL}/billing?canceled=true"

STRIPE_SYNC_EVENTS = {
    "checkout.session.completed",
    "customer.subscription.created",
    "customer.subscription.updated",
    "invoice.paid",
    "customer.subscription.deleted",
    "customer.subscription.paused",
    "customer.subscription.resumed",
    "customer.subscription.pending_update_applied",
    "customer.subscription.pending_update_expired",
    "customer.subscription.trial_will_end",
    "invoice.payment_failed",
    "invoice.payment_action_required",
    "invoice.upcoming",
    "invoice.marked_uncollectible",
    "invoice.payment_succeeded",
    "payment_intent.succeeded",
    "payment_intent.payment_failed",
    "payment_intent.canceled",
    # Fraud protection events - trigger immediate credit revocation
    "charge.refunded",
    "charge.dispute.created",
    "charge.dispute.closed",
    "charge.dispute.updated",
    "radar.early_fraud_warning.created",
}


# Validate critical environment variables at startup
def _validate_env_vars():
    """Validate that all critical environment variables are set."""
    warnings = []

    # Check Stripe configuration
    if not STRIPE_SECRET:
        warnings.append("STRIPE_SECRET_KEY is not set - payments will not work")

    if not STRIPE_WEBHOOK_SECRET:
        warnings.append("STRIPE_WEBHOOK_SECRET is not set - webhook verification will fail")

    # Check if all price IDs are configured
    missing_prices = []
    for plan, price_id in PLAN_PRICE_MAP.items():
        if not price_id:
            missing_prices.append(f"STRIPE_PRICE_{plan.upper()}")

    for plan, price_id in ADDON_PRICE_MAP.items():
        if not price_id:
            missing_prices.append(f"STRIPE_PRICE_ADDON_{plan.upper()}")

    if missing_prices:
        warnings.append(f"Missing Stripe price IDs: {', '.join(missing_prices)}")

    # Check Supabase configuration
    if not os.getenv("SUPABASE_URL"):
        warnings.append("SUPABASE_URL is not set - database operations will fail")

    if not os.getenv("SUPABASE_SERVICE_ROLE_KEY"):
        warnings.append("SUPABASE_SERVICE_ROLE_KEY is not set - database operations will fail")

    if not os.getenv("SUPABASE_JWT_SECRET"):
        warnings.append("SUPABASE_JWT_SECRET is not set - JWT validation will fail")

    # Log all warnings
    if warnings:
        print("=" * 80)
        print("ENVIRONMENT CONFIGURATION WARNINGS:")
        for warning in warnings:
            print(f"  ⚠️  {warning}")
            logging.warning(warning)
        print("=" * 80)
    else:
        logging.info("All critical environment variables are configured")


# Run validation on startup
_validate_env_vars()


def _fetch_profile(user_id: str, columns: str = "id,email,stripe_customer_id") -> Optional[Dict[str, Any]]:
    try:
        response = (
            supabase.table("profiles")
            .select(columns)
            .eq("id", user_id)
            .single()
            .execute()
        )
        if getattr(response, "data", None):
            return response.data
    except Exception as exc:  # pragma: no cover - supabase availability
        print(f"[Stripe] Failed to fetch profile for {user_id}: {exc}")
    return None


def _update_profile(user_id: str, payload: Dict[str, Any]) -> None:
    if not payload:
        return

    try:
        supabase.table("profiles").upsert(
            {**payload, "id": user_id}
        ).execute()
    except Exception as exc:  # pragma: no cover - supabase availability
        print(f"[Stripe] Failed to update profile {user_id}: {exc}")


def _find_user_by_customer(customer_id: str) -> Optional[str]:
    if not customer_id:
        return None

    try:
        response = (
            supabase.table("profiles")
            .select("id")
            .eq("stripe_customer_id", customer_id)
            .single()
            .execute()
        )
        if getattr(response, "data", None):
            return response.data.get("id")
    except Exception as exc:  # pragma: no cover - supabase availability
        print(f"[Stripe] Failed to map customer {customer_id} via Supabase: {exc}")

    return None


def ensure_stripe_customer_id(user_id: str, email: Optional[str] = None) -> Optional[str]:
    profile = _fetch_profile(user_id)
    if profile:
        stripe_customer_id = profile.get("stripe_customer_id")
        if stripe_customer_id:
            return stripe_customer_id
        email = email or profile.get("email")

    if not STRIPE_SECRET:
        return None

    customer_payload: Dict[str, Any] = {"metadata": {"user_id": user_id}}
    if email:
        customer_payload["email"] = email

    try:
        customer = stripe.Customer.create(**customer_payload)
        stripe_customer_id = customer.get("id")

        if stripe_customer_id:
            _update_profile(user_id, {"stripe_customer_id": stripe_customer_id})

        return stripe_customer_id
    except Exception as e:
        print(f"[ERROR] Failed to create Stripe customer for user {user_id}: {e}")
        return None


def sync_stripe_customer(customer_id: str, user_id: Optional[str] = None) -> Dict[str, Any]:
    if not STRIPE_SECRET or not customer_id:
        return {"status": "skipped", "reason": "missing_credentials"}

    if not user_id:
        user_id = _find_user_by_customer(customer_id)

    subscriptions = stripe.Subscription.list(
        customer=customer_id,
        limit=1,
        status="all",
        expand=["data.default_payment_method"],
    )

    if not subscriptions.data:
        snapshot: Dict[str, Any] = {"status": "none"}
    else:
        subscription = subscriptions.data[0]
        default_payment = subscription.get("default_payment_method")
        payment_method: Optional[Dict[str, Any]] = None
        if default_payment and not isinstance(default_payment, str):
            payment_method = {
                "brand": default_payment.get("card", {}).get("brand"),
                "last4": default_payment.get("card", {}).get("last4"),
            }

        snapshot = {
            "subscriptionId": subscription.get("id"),
            "status": subscription.get("status"),
            "priceId": subscription.get("items", {})
            .get("data", [{}])[0]
            .get("price", {})
            .get("id"),
            "currentPeriodEnd": subscription.get("current_period_end"),
            "currentPeriodStart": subscription.get("current_period_start"),
            "cancelAtPeriodEnd": subscription.get("cancel_at_period_end"),
            "paymentMethod": payment_method,
        }

    if user_id:
        profile_update: Dict[str, Any] = {
            "subscription_status": snapshot.get("status") or "inactive",
            "renewal_date": snapshot.get("currentPeriodEnd"),
            "stripe_subscription_id": snapshot.get("subscriptionId"),
            "stripe_price_id": snapshot.get("priceId"),
            "stripe_current_period_start": snapshot.get("currentPeriodStart"),
            "stripe_current_period_end": snapshot.get("currentPeriodEnd"),
            "stripe_cancel_at_period_end": snapshot.get("cancelAtPeriodEnd"),
            "stripe_payment_brand": snapshot.get("paymentMethod", {}).get("brand")
            if snapshot.get("paymentMethod")
            else None,
            "stripe_payment_last4": snapshot.get("paymentMethod", {}).get("last4")
            if snapshot.get("paymentMethod")
            else None,
        }

        plan_slug = PRICE_TO_PLAN.get(snapshot.get("priceId"))
        if plan_slug:
            profile_update["plan_type"] = plan_slug

        if snapshot.get("status") == "none":
            profile_update["subscription_status"] = "inactive"

        _update_profile(user_id, profile_update)

    return {
        "customerId": customer_id,
        "userId": user_id,
        "subscription": snapshot,
    }

from . import jobs

def start_worker():
    t = threading.Thread(target=jobs.worker_loop, daemon=True)
    t.start()


@app.on_event("startup")
async def startup_event():
    print("[Main] Running startup_event")
    start_worker()
    print("[Main] Worker started")


app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
def health():
    return {"status": "ok"}

# ---- Extract user from JWT ----
def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> AuthenticatedUser:
    token = credentials.credentials

    secret = os.getenv("SUPABASE_JWT_SECRET")
    if not secret:
        raise HTTPException(status_code=500, detail="SUPABASE_JWT_SECRET is not configured")

    supabase_url = os.getenv("SUPABASE_URL", "").rstrip("/")
    issuer = os.getenv("SUPABASE_JWT_ISS", f"{supabase_url}/auth/v1" if supabase_url else None)
    audience = os.getenv("SUPABASE_JWT_AUD", os.getenv("SUPABASE_JWT_AUDIENCE", "authenticated"))

    options = {"require": ["exp"]}
    if issuer:
        options["require"].append("iss")
    if audience:
        options["require"].append("aud")

    try:
        payload = jwt.decode(
            token,
            secret,
            algorithms=["HS256"],
            audience=audience if audience else None,
            issuer=issuer if issuer else None,
            options={**options, "verify_aud": bool(audience)},
        )
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except (InvalidAudienceError, InvalidIssuerError) as exc:
        raise HTTPException(status_code=401, detail=str(exc))
    except InvalidTokenError as exc:
        raise HTTPException(status_code=401, detail=f"Invalid token: {exc}")

    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Token missing subject")

    return AuthenticatedUser(user_id=user_id, claims=payload)

# ✅ Step 1 — Parse headers
def get_supabase():
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    return create_client(url, key)


class ParseRequest(BaseModel):
    file_path: str


def assert_user_owns_path(file_path: str, user_id: str) -> str:
    """Ensure the storage path belongs to the authenticated user."""
    if not file_path:
        raise HTTPException(status_code=403, detail="Forbidden")

    normalized_path = file_path.lstrip("/")
    expected_prefix = f"{user_id}/"

    if not normalized_path.startswith(expected_prefix):
        raise HTTPException(status_code=403, detail="Forbidden")

    return normalized_path


@app.post("/parse_headers")
async def parse_headers(
    payload: ParseRequest = Body(...),
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    try:
        file_path = payload.file_path
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path required")

        # Security: Validate file extension
        allowed_extensions = {'.csv', '.xlsx', '.xls', '.tsv'}
        file_ext = os.path.splitext(file_path.lower())[1]
        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid file type. Only CSV and Excel files are allowed."
            )

        file_path = assert_user_owns_path(file_path, current_user.user_id)

        supabase_client = get_supabase()
        temp_path = await stream_input_to_tempfile(supabase_client, file_path)
        print(f"[ParseHeaders] Streamed {file_path} to temporary file")

        try:
            if file_path.lower().endswith(".xlsx"):
                headers = extract_xlsx_headers(temp_path)
                row_count = count_xlsx_rows(temp_path)
            else:
                headers = extract_csv_headers(temp_path)
                row_count = count_csv_rows(temp_path)
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

        profile_res = (
            supabase_client.table("profiles")
            .select("credits_remaining, addon_credits")
            .eq("id", current_user.user_id)
            .single()
            .execute()
        )
        profile_data = getattr(profile_res, "data", None)
        if isinstance(profile_data, list):
            profile_data = profile_data[0] if profile_data else None
        credits_remaining = 0
        addon_credits = 0
        if isinstance(profile_data, dict):
            try:
                credits_remaining = int(profile_data.get("credits_remaining") or 0)
                addon_credits = int(profile_data.get("addon_credits") or 0)
            except (TypeError, ValueError):
                credits_remaining = 0
                addon_credits = 0

        total_credits = credits_remaining + addon_credits
        has_enough_credits = total_credits >= row_count
        missing_credits = max(0, row_count - total_credits)

        email_guess = ""
        email_variants = {"email", "emails", "e-mail", "e-mails", "mail", "mails"}
        for header in headers:
            normalized = header.strip().lower()
            normalized_simple = re.sub(r"[^a-z0-9]", "", normalized)
            if normalized in email_variants or normalized_simple in email_variants:
                email_guess = header
                break

        if not email_guess:
            pattern = re.compile(r"\b(e-?mail|mail)s?\b", re.IGNORECASE)
            for header in headers:
                if pattern.search(header):
                    email_guess = header
                    break

        print(
            f"[ParseHeaders] Parsed headers for {file_path}: {headers} — rows={row_count} "
            f"monthly={credits_remaining} addon={addon_credits} total={total_credits} enough={has_enough_credits} guess={email_guess!r}"
        )

        return {
            "headers": headers,
            "file_path": file_path,
            "row_count": row_count,
            "credits_remaining": total_credits,
            "has_enough_credits": has_enough_credits,
            "missing_credits": missing_credits,
            "email_header_guess": email_guess,
        }

    except HTTPException:
        raise
    except FileStreamingError as exc:
        print(f"[ParseHeaders] Streaming ERROR: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as e:
        print(f"[ParseHeaders] ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ✅ Step 2 — Create job
from pydantic import BaseModel

class ServiceComponents(BaseModel):
    core_offer: str
    key_differentiator: str
    cta: str
    include_fallback: Optional[bool] = None
    timeline: Optional[str] = None
    goal: Optional[str] = None
    fallback_action: Optional[str] = None


class JobRequest(BaseModel):
    file_path: str
    email_col: str
    service: Union[str, ServiceComponents]  # Accept either string (legacy) or structured components
    process_limit: Optional[int] = None


@app.get("/jobs")
def list_jobs(
    current_user: AuthenticatedUser = Depends(get_current_user),
    offset: int = 0,
    limit: int = 5
):
    supabase = get_supabase()
    user_id = current_user.user_id

    jobs_res = (
        supabase.table("jobs")
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .range(offset, offset + limit - 1)  # <-- pagination
        .execute()
    )
    jobs = jobs_res.data or []
    if not jobs:
        return []

    job_ids = [job["id"] for job in jobs]

    logs_res = (
        supabase.table("job_logs")
        .select("job_id, step, total, message")
        .in_("job_id", job_ids)
        .order("step", desc=True)
        .execute()
    )
    logs = logs_res.data or []
    latest_logs = {}
    for log in logs:
        job_id = log["job_id"]
        if job_id not in latest_logs:  # first (highest step) wins
            latest_logs[job_id] = log

    for job in jobs:
        last_log = latest_logs.get(job["id"])
        if last_log:
            step = last_log["step"]
            total = last_log["total"] or 1
            job["progress"] = int((step / total) * 100)
            job["message"] = last_log.get("message")
        else:
            job["progress"] = 0
            job["message"] = None

    return jobs



# ✅ Other routes unchanged (get_me, list_jobs, get_job, download, progress, checkout, webhook)
# ... keep all your code from [147†source] after this point


@app.get("/me")
def get_me(current_user: AuthenticatedUser = Depends(get_current_user)):
    try:
        user_id = current_user.user_id
        res = (
            supabase.table("profiles")
            .select("id, email, credits_remaining, addon_credits, max_credits, plan_type, subscription_status, renewal_date, welcome_reward_status, created_at, service_context")
            .eq("id", user_id)
            .single()
            .execute()
        )

        if not res.data:
            # bootstrap default profile
            profile = {
                "id": user_id,
                "email": "unknown@example.com",
                "credits_remaining": 500,
                "max_credits": 5000,
                "plan_type": "free",
                "subscription_status": "active",
                "renewal_date": (datetime.utcnow() + timedelta(days=30)).isoformat()
            }
            supabase.table("profiles").insert(profile).execute()
            res = (
                supabase.table("profiles")
                .select("*")
                .eq("id", user_id)
                .single()
                .execute()
            )

        return res.data
    except HTTPException:
        raise
    except FileStreamingError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/signup-reward")
def get_signup_reward(current_user: AuthenticatedUser = Depends(get_current_user)):
    """
    Get the current user's signup reward status and progress.
    Returns progress toward the 500 credit goal and time remaining.
    """
    try:
        user_id = current_user.user_id

        # Get signup reward record
        reward_res = (
            supabase.table("signup_rewards")
            .select("*")
            .eq("user_id", user_id)
            .single()
            .execute()
        )

        if not reward_res.data:
            # No reward record found - user may have signed up before this feature
            return {
                "has_reward": False,
                "message": "No signup reward available"
            }

        reward = reward_res.data
        status = reward.get("status", "active")
        credits_used = reward.get("credits_used", 0)
        credits_goal = reward.get("credits_goal", 500)
        reward_credits = reward.get("reward_credits", 500)
        deadline_at = reward.get("deadline_at")
        started_at = reward.get("started_at")
        unlocked_at = reward.get("unlocked_at")
        claimed_at = reward.get("claimed_at")

        # Check if active reward has expired (lazy expiration)
        if status == "active" and deadline_at:
            deadline = _parse_supabase_timestamp(deadline_at)
            if deadline:
                now = datetime.now(deadline.tzinfo) if deadline.tzinfo else datetime.utcnow()
                if now >= deadline:
                    # Mark as expired
                    supabase.table("signup_rewards").update({
                        "status": "expired"
                    }).eq("id", reward["id"]).execute()
                    status = "expired"

        # Calculate time remaining (for active status)
        time_remaining_seconds = None
        if status == "active" and deadline_at:
            deadline = _parse_supabase_timestamp(deadline_at)
            if deadline:
                now = datetime.now(deadline.tzinfo) if deadline.tzinfo else datetime.utcnow()
                remaining = (deadline - now).total_seconds()
                time_remaining_seconds = max(0, int(remaining))

        # Calculate progress percentage
        progress_percent = min(100, round((credits_used / credits_goal) * 100, 1)) if credits_goal > 0 else 0
        credits_remaining_to_goal = max(0, credits_goal - credits_used)

        return {
            "has_reward": True,
            "status": status,
            "credits_used": credits_used,
            "credits_goal": credits_goal,
            "credits_remaining_to_goal": credits_remaining_to_goal,
            "reward_credits": reward_credits,
            "progress_percent": progress_percent,
            "deadline_at": deadline_at,
            "started_at": started_at,
            "time_remaining_seconds": time_remaining_seconds,
            "unlocked_at": unlocked_at,
            "claimed_at": claimed_at
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"[Rewards] Get signup reward error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/signup-reward/claim")
def claim_signup_reward(current_user: AuthenticatedUser = Depends(get_current_user)):
    """
    Claim the signup reward (500 bonus credits) after unlocking it.
    User must have status 'unlocked' to claim.
    """
    try:
        user_id = current_user.user_id

        # Get signup reward record
        reward_res = (
            supabase.table("signup_rewards")
            .select("id, status, reward_credits")
            .eq("user_id", user_id)
            .single()
            .execute()
        )

        if not reward_res.data:
            raise HTTPException(status_code=404, detail="No signup reward found")

        reward = reward_res.data
        status = reward.get("status")
        reward_credits = reward.get("reward_credits", 500)

        if status == "claimed":
            raise HTTPException(status_code=400, detail="Reward already claimed")

        if status == "expired":
            raise HTTPException(status_code=400, detail="Reward has expired")

        if status != "unlocked":
            raise HTTPException(status_code=400, detail="Reward is not yet unlocked. Use 500 credits within 7 days to unlock.")

        # Get current credits
        profile_res = (
            supabase.table("profiles")
            .select("credits_remaining")
            .eq("id", user_id)
            .single()
            .execute()
        )

        if not profile_res.data:
            raise HTTPException(status_code=404, detail="User profile not found")

        current_credits = profile_res.data.get("credits_remaining", 0)

        # Update signup_rewards status to claimed
        supabase.table("signup_rewards").update({
            "status": "claimed",
            "claimed_at": datetime.utcnow().isoformat() + "Z"
        }).eq("id", reward["id"]).execute()

        # Add reward credits to profile
        supabase.table("profiles").update({
            "credits_remaining": current_credits + reward_credits
        }).eq("id", user_id).execute()

        # Add to ledger for audit trail
        supabase.table("ledger").insert({
            "user_id": user_id,
            "change": reward_credits,
            "amount": 0,
            "reason": "signup reward claimed",
            "ts": datetime.utcnow().isoformat()
        }).execute()

        print(f"[Rewards] User {user_id} claimed signup reward of {reward_credits} credits")

        return {
            "status": "success",
            "message": f"Successfully claimed {reward_credits} bonus credits!",
            "credits_added": reward_credits,
            "new_balance": current_credits + reward_credits
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"[Rewards] Claim signup reward error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/account/ledger")
def get_account_ledger(
    limit: int = 50,
    offset: int = 0,
    current_user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Get paginated ledger transactions for the current user
    """
    supabase = get_supabase()
    user_id = current_user.user_id

    # Get total count
    count_res = (
        supabase.table("ledger")
        .select("id", count="exact")
        .eq("user_id", user_id)
        .execute()
    )
    total = count_res.count or 0

    # Get paginated transactions
    ledger_res = (
        supabase.table("ledger")
        .select("*")
        .eq("user_id", user_id)
        .order("ts", desc=True)
        .range(offset, offset + limit - 1)
        .execute()
    )
    transactions = ledger_res.data or []

    return {
        "transactions": transactions,
        "total": total,
        "limit": limit,
        "offset": offset
    }

@app.get("/subscription/info")
def get_subscription_info(current_user: AuthenticatedUser = Depends(get_current_user)):
    """
    Get detailed subscription information for the current user
    """
    supabase = get_supabase()
    user_id = current_user.user_id

    # Get profile with subscription details
    profile_res = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .single()
        .execute()
    )

    if not profile_res.data:
        raise HTTPException(status_code=404, detail="Profile not found")

    profile = profile_res.data
    stripe_customer_id = profile.get("stripe_customer_id")

    subscription_info = {
        "plan_type": profile.get("plan_type", "free"),
        "subscription_status": profile.get("subscription_status"),
        "billing_frequency": profile.get("billing_frequency", "monthly"),
        "credits_remaining": profile.get("credits_remaining", 0),
        "addon_credits": profile.get("addon_credits", 0),
        "max_credits": CREDITS_MAP.get(profile.get("plan_type", "free"), 0),
        "cancel_at_period_end": False,
        "current_period_end": None,
        "upcoming_invoice": None,
    }

    # Fetch Stripe subscription details if customer exists
    if stripe_customer_id:
        try:
            # Get active subscriptions for this customer
            subscriptions = stripe.Subscription.list(
                customer=stripe_customer_id,
                status="active",
                limit=1
            )

            if subscriptions.data:
                sub = subscriptions.data[0]
                subscription_info["cancel_at_period_end"] = sub.get("cancel_at_period_end", False)
                subscription_info["current_period_end"] = sub.get("current_period_end")
                subscription_info["subscription_id"] = sub.get("id")

                # Check if there's a scheduled plan change
                schedule = sub.get("schedule")
                if schedule:
                    try:
                        schedule_obj = stripe.SubscriptionSchedule.retrieve(schedule)
                        phases = schedule_obj.get("phases", [])
                        if len(phases) > 1:
                            # There's a future phase (downgrade scheduled)
                            next_phase = phases[-1]
                            items = next_phase.get("items", [])
                            if items:
                                next_price_id = items[0].get("price")
                                next_plan = PRICE_TO_PLAN.get(next_price_id)
                                if next_plan:
                                    subscription_info["pending_plan_change"] = next_plan
                    except Exception as e:
                        print(f"Error fetching subscription schedule: {e}")

        except Exception as e:
            print(f"Error fetching Stripe subscription: {e}")

    # Also check profile for pending_downgrade (fallback if Stripe schedule lookup fails)
    pending_downgrade = profile.get("pending_downgrade")
    if pending_downgrade and "pending_plan_change" not in subscription_info:
        subscription_info["pending_plan_change"] = pending_downgrade
        # Add effective date if available
        effective_date = profile.get("pending_downgrade_effective_date")
        if effective_date:
            subscription_info["pending_plan_change_date"] = effective_date

    return subscription_info


@app.get("/billing/invoices")
def list_invoices(
    limit: int = 20,
    current_user: AuthenticatedUser = Depends(get_current_user)
):
    """
    List invoices for the current user from Stripe.
    Returns invoice number, amount, status, date, and PDF download URL.
    """
    user_id = current_user.user_id
    
    # Get profile with stripe_customer_id
    profile = _fetch_profile(user_id, columns="id,stripe_customer_id")
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    
    stripe_customer_id = profile.get("stripe_customer_id")
    if not stripe_customer_id:
        # No Stripe customer = no invoices
        return {"invoices": []}
    
    try:
        # Fetch invoices from Stripe
        invoices = stripe.Invoice.list(
            customer=stripe_customer_id,
            limit=limit,
        )
        
        invoice_list = []
        for inv in invoices.data:
            # Only include paid/finalized invoices
            if inv.status not in ["paid", "open", "uncollectible"]:
                continue
                
            invoice_list.append({
                "id": inv.id,
                "number": inv.number or inv.id[:12],
                "amount": (inv.amount_paid or 0) / 100,  # Convert cents to dollars
                "status": inv.status,
                "pdf_url": inv.invoice_pdf,
                "hosted_url": inv.hosted_invoice_url,
                "date": inv.created,  # Unix timestamp
                "description": inv.description or _get_invoice_description(inv),
            })
        
        return {"invoices": invoice_list}
        
    except stripe.error.StripeError as e:
        print(f"[INVOICES] Stripe error: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch invoices")


def _get_invoice_description(invoice) -> str:
    """Generate a description from invoice line items"""
    if invoice.lines and invoice.lines.data:
        first_line = invoice.lines.data[0]
        if first_line.description:
            return first_line.description
        if first_line.price and first_line.price.nickname:
            return first_line.price.nickname
    return "Invoice"


@app.get("/subscription/upgrade/preview")
def preview_upgrade(
    plan: str,
    current_user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Preview an upgrade - returns exact bonus credits, total credits, and price
    without performing the actual upgrade. Used to show accurate info in the UI.
    """
    from datetime import datetime
    
    supabase = get_supabase()
    user_id = current_user.user_id

    # Extract base plan name
    base_plan = plan.replace("_annual", "").replace("_monthly", "")
    
    if base_plan not in CREDITS_MAP:
        raise HTTPException(status_code=400, detail="Invalid plan")

    # Get profile
    profile_res = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .single()
        .execute()
    )
    
    if not profile_res.data:
        raise HTTPException(status_code=404, detail="Profile not found")

    profile = profile_res.data
    current_plan = profile.get("plan_type", "free")
    stripe_subscription_id = profile.get("stripe_subscription_id")
    
    base_current_plan = current_plan.replace("_annual", "").replace("_monthly", "")
    base_new_plan = plan.replace("_annual", "").replace("_monthly", "")
    
    # Get plan prices and credits
    base_monthly_price = PLAN_PRICES.get(base_new_plan, 0)
    if plan.endswith("_annual"):
        new_plan_price = int(base_monthly_price * 12 * 0.8)
        new_plan_credits = CREDITS_MAP.get(base_new_plan, 0) * 12
    else:
        new_plan_price = base_monthly_price
        new_plan_credits = CREDITS_MAP.get(base_new_plan, 0)
    
    old_plan_credits = CREDITS_MAP.get(base_current_plan, 0)
    old_plan_price = PLAN_PRICES.get(base_current_plan, 0)
    current_credits = profile.get("credits_remaining", 0)
    
    # Calculate bonus credits (same logic as upgrade endpoint)
    bonus_credits = 0
    if stripe_subscription_id and old_plan_price > 0:
        try:
            subscription = stripe.Subscription.retrieve(stripe_subscription_id)
            sub_dict = dict(subscription)
            period_start = sub_dict.get('current_period_start')
            period_end = sub_dict.get('current_period_end')
            
            if not period_start or not period_end:
                start_date = sub_dict.get('start_date') or sub_dict.get('created')
                if start_date:
                    now = datetime.utcnow().timestamp()
                    days_since_start = (now - start_date) / 86400
                    day_in_cycle = days_since_start % 30
                    remaining_days = 30 - day_in_cycle
                    unused_ratio = remaining_days / 30
                    bonus_credits = int(old_plan_credits * unused_ratio)
                    bonus_credits = min(bonus_credits, new_plan_credits)
            elif old_plan_price > 0:
                now = datetime.utcnow().timestamp()
                total_days = (period_end - period_start) / 86400
                remaining_days = max(0, (period_end - now) / 86400)
                if total_days > 0:
                    unused_ratio = remaining_days / total_days
                    bonus_credits = int(old_plan_credits * unused_ratio)
                    bonus_credits = min(bonus_credits, new_plan_credits)
        except Exception as e:
            print(f"[PREVIEW] Bonus calculation error: {e}")

    final_credits = new_plan_credits + bonus_credits

    return {
        "current_plan": current_plan,
        "new_plan": plan,
        "current_credits": current_credits,
        "new_plan_credits": new_plan_credits,
        "bonus_credits": bonus_credits,
        "total_credits": final_credits,
        "price": new_plan_price,
        "billing_cycle": "annual" if plan.endswith("_annual") else "monthly"
    }

@app.post("/subscription/upgrade")
def upgrade_subscription(
    plan: str,
    current_user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Upgrade subscription to a higher plan.
    
    FLOW:
    1. Validate plan and subscription
    2. Check payment method exists FIRST
    3. Check idempotency (not already processing)
    4. Set processing flag
    5. Create and PAY invoice FIRST (before modifying subscription)
    6. ONLY if payment succeeds, modify subscription
    7. Update profile only after payment + subscription both succeed
    8. Clear processing flag
    """
    from datetime import datetime
    
    supabase = get_supabase()
    user_id = current_user.user_id

    # Extract base plan name for validation (remove _annual/_monthly suffix)
    base_plan = plan.replace("_annual", "").replace("_monthly", "")
    
    # Validate plan
    if base_plan not in CREDITS_MAP:
        raise HTTPException(status_code=400, detail="Invalid plan")

    # Get current profile
    profile_res = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .single()
        .execute()
    )

    if not profile_res.data:
        raise HTTPException(status_code=404, detail="Profile not found")

    profile = profile_res.data
    current_plan = profile.get("plan_type", "free")
    stripe_subscription_id = profile.get("stripe_subscription_id")
    stripe_customer_id = profile.get("stripe_customer_id")

    # ============================================================
    # STEP 1: IDEMPOTENCY CHECK - prevent double-processing
    # ============================================================
    if profile.get("upgrade_in_progress"):
        raise HTTPException(
            status_code=409, 
            detail="An upgrade is already being processed. Please wait a moment and try again."
        )

    if not stripe_subscription_id:
        raise HTTPException(status_code=400, detail="No active subscription found. Please subscribe to a plan first.")

    # ============================================================
    # STEP 2: VALIDATE PAYMENT METHOD FIRST (before any mutations)
    # ============================================================
    customer = stripe.Customer.retrieve(stripe_customer_id)
    default_pm = customer.get("invoice_settings", {}).get("default_payment_method")
    if not default_pm:
        # Fallback: get first payment method attached to customer
        payment_methods = stripe.PaymentMethod.list(customer=stripe_customer_id, type="card", limit=1)
        if payment_methods.data:
            default_pm = payment_methods.data[0].id

    if not default_pm:
        raise HTTPException(
            status_code=400, 
            detail="No payment method on file. Please add a payment method before upgrading."
        )

    # ============================================================
    # STEP 3: VALIDATE SUBSCRIPTION STATUS
    # ============================================================
    try:
        stripe_sub = stripe.Subscription.retrieve(stripe_subscription_id)
        sub_status = stripe_sub.get("status")
        if sub_status in ["canceled", "incomplete_expired"]:
            supabase.table("profiles").update({
                "stripe_subscription_id": None,
                "subscription_status": "canceled",
            }).eq("id", user_id).execute()
            raise HTTPException(
                status_code=400, 
                detail="Your subscription has been canceled. Please subscribe to a new plan from the billing page."
            )
        elif sub_status not in ["active", "trialing"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot upgrade: subscription status is '{sub_status}'. Please contact support."
            )
    except stripe.error.InvalidRequestError:
        supabase.table("profiles").update({
            "stripe_subscription_id": None,
        }).eq("id", user_id).execute()
        raise HTTPException(
            status_code=400, 
            detail="Subscription not found. Please subscribe to a new plan."
        )

    # ============================================================
    # STEP 4: VALIDATE UPGRADE DIRECTION
    # ============================================================
    base_current_plan = current_plan.replace("_annual", "").replace("_monthly", "")
    base_new_plan = plan.replace("_annual", "").replace("_monthly", "")
    is_annual_switch = plan.endswith("_annual") and not current_plan.endswith("_annual")
    
    # Block annual→monthly switch
    if current_plan.endswith("_annual") and not plan.endswith("_annual"):
        raise HTTPException(
            status_code=400,
            detail="Please cancel your annual plan to switch to monthly billing."
        )

    current_plan_credits = CREDITS_MAP.get(base_current_plan, 0)
    new_plan_credits = CREDITS_MAP.get(base_new_plan, 0)

    # Block downgrades
    if new_plan_credits < current_plan_credits:
        raise HTTPException(status_code=400, detail="Downgrades are not supported.")
    
    # Block same plan (unless annual switch)
    if new_plan_credits == current_plan_credits and not is_annual_switch:
        raise HTTPException(status_code=400, detail="You're already on this plan.")

    # Get new price ID
    if plan.endswith("_annual"):
        new_price_id = PLAN_PRICE_MAP_ANNUAL.get(plan)
    else:
        new_price_id = PLAN_PRICE_MAP.get(plan)
    
    if not new_price_id:
        raise HTTPException(status_code=400, detail="Plan price not configured")

    # Calculate price
    base_monthly_price = PLAN_PRICES.get(base_new_plan, 0)
    if plan.endswith("_annual"):
        new_plan_price = int(base_monthly_price * 12 * 0.8)
    else:
        new_plan_price = base_monthly_price
    
    old_plan_price = PLAN_PRICES.get(base_current_plan, 0)
    old_plan_credits = CREDITS_MAP.get(base_current_plan, 0)
    
    # For annual, multiply credits by 12
    if plan.endswith("_annual"):
        new_plan_credits = new_plan_credits * 12

    # ============================================================
    # STEP 5: SET PROCESSING FLAG (idempotency)
    # ============================================================
    supabase.table("profiles").update({
        "upgrade_in_progress": True
    }).eq("id", user_id).execute()
    print(f"[UPGRADE] Processing flag set for user {user_id}")

    try:
        # Get subscription item
        items = stripe.SubscriptionItem.list(subscription=stripe_subscription_id, limit=1)
        if not items.data:
            raise HTTPException(status_code=400, detail="No subscription items found")
        item_id = items.data[0].id

        # Calculate bonus credits
        bonus_credits = 0
        try:
            subscription = stripe.Subscription.retrieve(stripe_subscription_id)
            sub_dict = dict(subscription)
            period_start = sub_dict.get('current_period_start')
            period_end = sub_dict.get('current_period_end')
            
            if not period_start or not period_end:
                start_date = sub_dict.get('start_date') or sub_dict.get('created')
                if start_date and old_plan_price > 0:
                    now = datetime.utcnow().timestamp()
                    days_since_start = (now - start_date) / 86400
                    day_in_cycle = days_since_start % 30
                    remaining_days = 30 - day_in_cycle
                    unused_ratio = remaining_days / 30
                    bonus_credits = int(old_plan_credits * unused_ratio)
                    bonus_credits = min(bonus_credits, new_plan_credits)
                    print(f"[UPGRADE] Bonus (via start_date): {remaining_days:.1f}/30 days = {bonus_credits} credits")
            elif old_plan_price > 0:
                now = datetime.utcnow().timestamp()
                total_days = (period_end - period_start) / 86400
                remaining_days = max(0, (period_end - now) / 86400)
                if total_days > 0:
                    unused_ratio = remaining_days / total_days
                    bonus_credits = int(old_plan_credits * unused_ratio)
                    bonus_credits = min(bonus_credits, new_plan_credits)
                    print(f"[UPGRADE] Bonus: {remaining_days:.1f}/{total_days:.1f} days = {bonus_credits} credits")
        except Exception as e:
            print(f"[UPGRADE] Bonus calculation error: {e}")

        final_credits = new_plan_credits + bonus_credits
        print(f"[UPGRADE] {current_plan}→{plan}: base={new_plan_credits}, bonus={bonus_credits}, total={final_credits}")

        # ============================================================
        # STEP 6: CHARGE FIRST (before modifying subscription!)
        # ============================================================
        invoice = stripe.Invoice.create(
            customer=stripe_customer_id,
            auto_advance=False,
        )
        
        stripe.InvoiceItem.create(
            customer=stripe_customer_id,
            invoice=invoice.id,
            amount=int(new_plan_price * 100),
            currency="usd",
            description=f"Upgrade to {plan.capitalize()} Plan"
        )
        
        invoice = stripe.Invoice.finalize_invoice(invoice.id)
        print(f"[UPGRADE] Invoice {invoice.id} created and finalized, amount: ${new_plan_price}")
        
        # Try to pay the invoice
        try:
            invoice = stripe.Invoice.pay(invoice.id, payment_method=default_pm)
            print(f"[UPGRADE] Invoice {invoice.id} payment attempted, status: {invoice.status}")
        except stripe.error.StripeError as payment_error:
            # Payment failed - void invoice, clear flag, return error
            print(f"[UPGRADE] Payment failed: {payment_error}")
            try:
                stripe.Invoice.void_invoice(invoice.id)
            except:
                pass  # Invoice might already be voided or uncollectible
            
            # Clear processing flag
            supabase.table("profiles").update({
                "upgrade_in_progress": False
            }).eq("id", user_id).execute()
            
            raise HTTPException(
                status_code=400, 
                detail=f"Payment failed: {str(payment_error)}"
            )
        
        # Verify payment succeeded
        invoice = stripe.Invoice.retrieve(invoice.id)
        if invoice.status != "paid":
            print(f"[UPGRADE] Invoice not paid, status: {invoice.status}")
            # Try to void the invoice
            try:
                stripe.Invoice.void_invoice(invoice.id)
            except:
                pass
            
            # Clear processing flag
            supabase.table("profiles").update({
                "upgrade_in_progress": False
            }).eq("id", user_id).execute()
            
            raise HTTPException(
                status_code=400, 
                detail=f"Payment was not successful. Please try again or use a different payment method."
            )
        
        print(f"[UPGRADE] Payment successful! Invoice {invoice.id} paid.")

        # ============================================================
        # STEP 7: PAYMENT SUCCEEDED - NOW MODIFY SUBSCRIPTION
        # ============================================================
        try:
            stripe.Subscription.modify(
                stripe_subscription_id,
                items=[{"id": item_id, "price": new_price_id}],
                proration_behavior="none",
                metadata={
                    "user_id": user_id,
                    "upgrade_from": current_plan,
                    "upgrade_to": plan,
                }
            )
            print(f"[UPGRADE] Subscription modified: {current_plan}→{plan}")
        except stripe.error.StripeError as sub_error:
            # Subscription modification failed after payment - REFUND
            print(f"[UPGRADE] Subscription modify failed after payment: {sub_error}")
            try:
                if invoice.payment_intent:
                    stripe.Refund.create(payment_intent=invoice.payment_intent)
                    print(f"[UPGRADE] Refunded payment_intent {invoice.payment_intent}")
            except Exception as refund_error:
                print(f"[UPGRADE] Refund failed: {refund_error}")
            
            # Clear processing flag
            supabase.table("profiles").update({
                "upgrade_in_progress": False
            }).eq("id", user_id).execute()
            
            raise HTTPException(
                status_code=400, 
                detail=f"Failed to update subscription after payment. Your payment has been refunded. Error: {str(sub_error)}"
            )

        # ============================================================
        # STEP 8: BOTH PAYMENT AND SUBSCRIPTION SUCCEEDED - UPDATE PROFILE
        # ============================================================
        new_billing_frequency = "annual" if plan.endswith("_annual") else "monthly"
        stored_plan_type = base_new_plan
        
        supabase.table("profiles").update({
            "plan_type": stored_plan_type,
            "billing_frequency": new_billing_frequency,
            "credits_remaining": final_credits,
            "subscription_status": "active",
            "pending_downgrade": None,  # Clear any pending downgrade
            "pending_downgrade_effective_date": None,
            "upgrade_in_progress": False,  # Clear processing flag
        }).eq("id", user_id).execute()
        print(f"[UPGRADE] Profile updated: plan={stored_plan_type}, credits={final_credits}, frequency={new_billing_frequency}")

        # Add ledger entry
        supabase.table("ledger").insert({
            "user_id": user_id,
            "change": final_credits,
            "amount": new_plan_price,
            "reason": f"plan upgrade - {current_plan} to {plan} (+{bonus_credits} bonus)",
            "ts": datetime.utcnow().isoformat(),
        }).execute()

        print(f"[UPGRADE] SUCCESS: Charged ${new_plan_price}, credits={final_credits}")

        return {
            "status": "success",
            "message": f"Upgraded to {plan} plan! Charged ${new_plan_price}.",
            "new_plan": plan,
            "credits": final_credits,
            "bonus_credits": bonus_credits,
            "amount_charged": new_plan_price,
        }

    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except stripe.error.StripeError as e:
        print(f"[UPGRADE] Stripe error: {e}")
        # Clear processing flag on any Stripe error
        supabase.table("profiles").update({
            "upgrade_in_progress": False
        }).eq("id", user_id).execute()
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"[UPGRADE] Unexpected error: {e}")
        # Clear processing flag on any error
        supabase.table("profiles").update({
            "upgrade_in_progress": False
        }).eq("id", user_id).execute()
        raise HTTPException(status_code=500, detail="An unexpected error occurred. Please try again.")

@app.post("/subscription/downgrade")
def downgrade_subscription(
    plan: str,
    current_user: AuthenticatedUser = Depends(get_current_user)
):
    """
    Schedule a downgrade to take effect at next billing cycle.
    User keeps current plan and credits until then.
    Uses Stripe's proration_behavior="none" for seamless transition.
    """
    from datetime import datetime
    
    supabase = get_supabase()
    user_id = current_user.user_id

    # Validate plan
    if plan not in CREDITS_MAP:
        raise HTTPException(status_code=400, detail="Invalid plan")

    # Get current profile
    profile_res = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .single()
        .execute()
    )

    if not profile_res.data:
        raise HTTPException(status_code=404, detail="Profile not found")

    profile = profile_res.data
    current_plan = profile.get("plan_type", "free")
    stripe_subscription_id = profile.get("stripe_subscription_id")

    if not stripe_subscription_id:
        raise HTTPException(status_code=400, detail="No active subscription found. Please subscribe to a plan first.")

    # Verify subscription is active in Stripe before attempting downgrade
    try:
        stripe_sub = stripe.Subscription.retrieve(stripe_subscription_id)
        sub_status = stripe_sub.get("status")
        if sub_status in ["canceled", "incomplete_expired"]:
            supabase.table("profiles").update({
                "stripe_subscription_id": None,
                "subscription_status": "canceled",
            }).eq("id", user_id).execute()
            raise HTTPException(
                status_code=400, 
                detail="Your subscription has been canceled. Please subscribe to a new plan."
            )
        elif sub_status not in ["active", "trialing"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot downgrade: subscription status is '{sub_status}'."
            )
    except stripe.error.InvalidRequestError:
        supabase.table("profiles").update({
            "stripe_subscription_id": None,
        }).eq("id", user_id).execute()
        raise HTTPException(status_code=400, detail="Subscription not found. Please subscribe to a new plan.")

    # Verify this is actually a downgrade
    current_plan_credits = CREDITS_MAP.get(current_plan, 0)
    new_plan_credits = CREDITS_MAP.get(plan, 0)

    if new_plan_credits >= current_plan_credits:
        raise HTTPException(status_code=400, detail="This is not a downgrade. Use /subscription/upgrade instead.")

    # Get new price ID
    new_price_id = PLAN_PRICE_MAP.get(plan)
    if not new_price_id:
        raise HTTPException(status_code=400, detail="Plan price not configured")

    try:
        # Get current subscription details to know when downgrade takes effect
        subscription = stripe.Subscription.retrieve(stripe_subscription_id)
        current_period_end = subscription.get("current_period_end")
        
        print(f"[DOWNGRADE] Scheduling: {current_plan}→{plan}, user={user_id}, effective at period_end={current_period_end}")

        # DO NOT modify the Stripe subscription!
        # Just store the pending downgrade - it will be applied in invoice.paid webhook
        # when the subscription renews to the new billing period
        
        supabase.table("profiles").update({
            "pending_downgrade": plan,
            "pending_downgrade_effective_date": current_period_end,
        }).eq("id", user_id).execute()

        # Add ledger entry for audit
        supabase.table("ledger").insert({
            "user_id": user_id,
            "change": 0,
            "amount": 0,
            "reason": f"downgrade scheduled - {current_plan} to {plan} (effective at period end: {current_period_end})",
            "ts": datetime.utcnow().isoformat(),
        }).execute()
        
        print(f"[DOWNGRADE] Scheduled {current_plan}→{plan} for user={user_id}")

        return {
            "status": "success",
            "message": f"Downgrade to {plan} scheduled. Takes effect at next billing cycle.",
            "current_plan": current_plan,
            "new_plan": plan,
            "effective_date": current_period_end,
        }

    except stripe.error.StripeError as e:
        print(f"[DOWNGRADE] Stripe error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/subscription/cancel")
def cancel_subscription(current_user: AuthenticatedUser = Depends(get_current_user)):
    """
    Cancel subscription at period end (user keeps access until then)
    """
    supabase = get_supabase()
    user_id = current_user.user_id

    # Get current profile
    profile_res = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .single()
        .execute()
    )

    if not profile_res.data:
        raise HTTPException(status_code=404, detail="Profile not found")

    profile = profile_res.data
    stripe_customer_id = profile.get("stripe_customer_id")

    if not stripe_customer_id:
        raise HTTPException(status_code=400, detail="No active subscription found")

    try:
        # Get current subscription
        subscriptions = stripe.Subscription.list(
            customer=stripe_customer_id,
            status="active",
            limit=1
        )

        if not subscriptions.data:
            raise HTTPException(status_code=400, detail="No active subscription found")

        subscription = subscriptions.data[0]
        subscription_id = subscription.id

        # Cancel at period end
        updated_subscription = stripe.Subscription.modify(
            subscription_id,
            cancel_at_period_end=True
        )

        return {
            "status": "success",
            "message": "Subscription will be canceled at the end of the current period",
            "subscription_id": updated_subscription.id,
            "cancel_at": updated_subscription.current_period_end
        }

    except stripe.error.StripeError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/subscription/reactivate")
def reactivate_subscription(current_user: AuthenticatedUser = Depends(get_current_user)):
    """
    Reactivate a subscription that was scheduled for cancellation
    """
    supabase = get_supabase()
    user_id = current_user.user_id

    # Get current profile
    profile_res = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .single()
        .execute()
    )

    if not profile_res.data:
        raise HTTPException(status_code=404, detail="Profile not found")

    profile = profile_res.data
    stripe_customer_id = profile.get("stripe_customer_id")

    if not stripe_customer_id:
        raise HTTPException(status_code=400, detail="No subscription found")

    try:
        # Get current subscription
        subscriptions = stripe.Subscription.list(
            customer=stripe_customer_id,
            limit=1
        )

        if not subscriptions.data:
            raise HTTPException(status_code=400, detail="No subscription found")

        subscription = subscriptions.data[0]
        subscription_id = subscription.id

        if not subscription.cancel_at_period_end:
            raise HTTPException(status_code=400, detail="Subscription is not scheduled for cancellation")

        # Reactivate by removing cancel_at_period_end
        updated_subscription = stripe.Subscription.modify(
            subscription_id,
            cancel_at_period_end=False
        )

        return {
            "status": "success",
            "message": "Subscription reactivated successfully",
            "subscription_id": updated_subscription.id
        }

    except stripe.error.StripeError as e:
        raise HTTPException(status_code=400, detail=str(e))




def _parse_profile_response(profile_res) -> Optional[Dict[str, Any]]:
    profile_data = getattr(profile_res, "data", None)
    if isinstance(profile_data, list):
        return profile_data[0] if profile_data else None
    if isinstance(profile_data, dict):
        return profile_data
    return None


def _reserve_credits_for_job(
    supabase_client,
    user_id: str,
    row_count: int,
    job_id: str,
    file_path: str,
) -> Dict[str, int]:
    max_attempts = 5
    for attempt in range(max_attempts):
        profile_res = (
            supabase_client.table("profiles")
            .select("credits_remaining, addon_credits")
            .eq("id", user_id)
            .limit(1)
            .execute()
        )
        profile = _parse_profile_response(profile_res) or {}
        try:
            credits_remaining = int(profile.get("credits_remaining") or 0)
            addon_credits = int(profile.get("addon_credits") or 0)
        except (TypeError, ValueError):
            credits_remaining = 0
            addon_credits = 0

        total_credits = credits_remaining + addon_credits

        # Check subscription status
        subscription_status = profile.get("subscription_status", "active")
        if subscription_status in ["past_due", "unpaid"]:
             raise HTTPException(
                status_code=402,
                detail={
                    "error": "subscription_past_due",
                    "message": "Your subscription is past due. Please update your payment method to continue.",
                    "status": subscription_status
                },
            )

        if total_credits < row_count:
            missing = row_count - total_credits
            print(
                f"[Credits] User {user_id} lacks {missing} credits for file {file_path}"
            )
            raise HTTPException(
                status_code=402,
                detail={
                    "error": "insufficient_credits",
                    "row_count": row_count,
                    "credits_remaining": total_credits,
                    "missing_credits": missing,
                },
            )

        # Two-bucket deduction: use monthly credits first, then add-ons
        if credits_remaining >= row_count:
            # Use only monthly credits
            new_monthly = credits_remaining - row_count
            new_addon = addon_credits
            monthly_deducted = row_count
            addon_deducted = 0
            deduction_source = "monthly"
        else:
            # Use all monthly + some add-on credits
            remaining_needed = row_count - credits_remaining
            new_monthly = 0
            new_addon = addon_credits - remaining_needed
            monthly_deducted = credits_remaining
            addon_deducted = remaining_needed
            deduction_source = "monthly+addon"

        update_res = (
            supabase_client.table("profiles")
            .update({
                "credits_remaining": new_monthly,
                "addon_credits": new_addon
            })
            .eq("id", user_id)
            .eq("credits_remaining", credits_remaining)
            .eq("addon_credits", addon_credits)
            .execute()
        )

        if getattr(update_res, "error", None):
            raise HTTPException(status_code=500, detail="Unable to reserve credits")

        updated_rows = getattr(update_res, "data", None) or []
        if updated_rows:
            ledger_payload = {
                "user_id": user_id,
                "change": -row_count,
                "amount": 0.0,
                "reason": f"job deduction: {job_id} ({deduction_source})",
                "ts": datetime.utcnow().isoformat(),
            }
            ledger_res = (
                supabase_client.table("ledger").insert(ledger_payload).execute()
            )
            if getattr(ledger_res, "error", None):
                supabase_client.table("profiles").update({
                    "credits_remaining": credits_remaining,
                    "addon_credits": addon_credits
                }).eq("id", user_id).execute()
                raise HTTPException(
                    status_code=500, detail="Unable to record credit deduction"
                )

            # Update signup reward progress after successful credit deduction
            _update_signup_reward_after_deduction(supabase_client, user_id, row_count)

            return {
                "previous_balance": credits_remaining + addon_credits,
                "new_balance": new_monthly + new_addon,
                "cost": row_count,
                "monthly_deducted": monthly_deducted,
                "addon_deducted": addon_deducted,
            }

        time.sleep(0.1)

    raise HTTPException(status_code=500, detail="Unable to reserve credits")


def _parse_supabase_timestamp(ts_str: str) -> datetime:
    """
    Parse timestamp string from Supabase with variable microsecond precision.
    Handles formats like '2025-12-11T22:48:34.19879+00:00' or '2025-12-11 22:48:34+00'
    """
    if not ts_str:
        return None
    try:
        # Normalize: replace space with T, handle Z suffix
        normalized = ts_str.replace(' ', 'T')
        if normalized.endswith('Z'):
            normalized = normalized[:-1] + '+00:00'

        # Handle +00 without :00
        if normalized.endswith('+00') and not normalized.endswith('+00:00'):
            normalized += ':00'

        # Try parsing directly
        try:
            return datetime.fromisoformat(normalized)
        except ValueError:
            pass

        # If that fails due to microsecond precision, normalize microseconds
        import re
        # Match microseconds and pad/truncate to 6 digits
        def fix_microseconds(match):
            micro = match.group(1)
            if len(micro) < 6:
                micro = micro.ljust(6, '0')
            elif len(micro) > 6:
                micro = micro[:6]
            return '.' + micro

        normalized = re.sub(r'\.(\d+)', fix_microseconds, normalized)
        return datetime.fromisoformat(normalized)
    except Exception as e:
        print(f"[Rewards] Failed to parse timestamp '{ts_str}': {e}")
        return None


def _update_signup_reward_after_deduction(supabase_client, user_id: str, credits_deducted: int):
    """
    Update signup_rewards table after credits are deducted.
    Called from _reserve_credits_for_job in the API.
    """
    if credits_deducted <= 0:
        return

    try:
        # Get current signup reward record
        reward_res = (
            supabase_client.table("signup_rewards")
            .select("id, credits_used, credits_goal, deadline_at, status")
            .eq("user_id", user_id)
            .single()
            .execute()
        )

        if not reward_res.data:
            print(f"[Rewards] No signup_rewards record found for user {user_id}")
            return

        reward = reward_res.data
        current_status = reward.get("status")

        # Only update if status is 'active'
        if current_status != "active":
            print(f"[Rewards] Signup reward for user {user_id} is already {current_status}, skipping update")
            return

        # Check if deadline has passed
        deadline_str = reward.get("deadline_at")
        if deadline_str:
            deadline = _parse_supabase_timestamp(deadline_str)
            if deadline:
                now = datetime.now(deadline.tzinfo) if deadline.tzinfo else datetime.utcnow()
                if now >= deadline:
                    supabase_client.table("signup_rewards").update({
                        "status": "expired"
                    }).eq("id", reward["id"]).execute()
                    print(f"[Rewards] Signup reward expired for user {user_id}")
                    return

        # Increment credits_used
        current_used = reward.get("credits_used", 0)
        credits_goal = reward.get("credits_goal", 500)
        new_used = current_used + credits_deducted

        # Check if goal is reached
        if new_used >= credits_goal:
            supabase_client.table("signup_rewards").update({
                "credits_used": new_used,
                "status": "unlocked",
                "unlocked_at": datetime.utcnow().isoformat() + "Z"
            }).eq("id", reward["id"]).execute()
            print(f"[Rewards] Unlocked signup reward for user {user_id} (used {new_used}/{credits_goal} credits)")
        else:
            supabase_client.table("signup_rewards").update({
                "credits_used": new_used
            }).eq("id", reward["id"]).execute()
            print(f"[Rewards] Updated signup reward progress for user {user_id}: {new_used}/{credits_goal}")

    except Exception as e:
        # Non-critical: don't fail the credit deduction if reward update fails
        print(f"[Rewards] Error updating signup reward for user {user_id}: {e}")


def _rollback_credit_reservation(
    supabase_client,
    user_id: str,
    reservation: Dict[str, int],
    job_id: str,
):
    """
    Rollback credit reservation by restoring credits to their original buckets.
    Uses the breakdown information to restore monthly and addon credits separately.
    """
    try:
        monthly_deducted = reservation.get("monthly_deducted", 0)
        addon_deducted = reservation.get("addon_deducted", 0)

        # Get current state for CAS update
        profile_res = (
            supabase_client.table("profiles")
            .select("credits_remaining, addon_credits")
            .eq("id", user_id)
            .limit(1)
            .execute()
        )

        if not profile_res.data:
            print(f"[Credits] Failed to rollback: profile not found for user {user_id}")
            return

        profile = profile_res.data[0]
        current_monthly = int(profile.get("credits_remaining") or 0)
        current_addon = int(profile.get("addon_credits") or 0)

        # Restore credits to both buckets
        new_monthly = current_monthly + monthly_deducted
        new_addon = current_addon + addon_deducted

        supabase_client.table("profiles").update({
            "credits_remaining": new_monthly,
            "addon_credits": new_addon
        }).eq("id", user_id).eq(
            "credits_remaining", current_monthly
        ).eq(
            "addon_credits", current_addon
        ).execute()

        supabase_client.table("ledger").insert(
            {
                "user_id": user_id,
                "change": reservation["cost"],
                "amount": 0.0,
                "reason": f"job rollback: {job_id} (monthly: +{monthly_deducted}, addon: +{addon_deducted})",
                "ts": datetime.utcnow().isoformat(),
            }
        ).execute()

        print(
            f"[Credits] Rolled back {reservation['cost']} credits for job {job_id} "
            f"(monthly: +{monthly_deducted}, addon: +{addon_deducted})"
        )
    except Exception as exc:  # pragma: no cover - logging for rollback issues
        print(f"[Credits] Failed to rollback reservation for job {job_id}: {exc}")


@app.post("/jobs")
async def create_job(
    req: JobRequest,
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    supabase = get_supabase()
    reservation: Optional[Dict[str, int]] = None
    lock = None
    acquired = False

    try:
        file_path = assert_user_owns_path(req.file_path, current_user.user_id)

        # Security: Validate file extension
        allowed_extensions = {'.csv', '.xlsx', '.xls', '.tsv'}
        file_ext = os.path.splitext(file_path.lower())[1]
        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400, 
                detail="Invalid file type. Only CSV and Excel files are allowed."
            )

        email_col = (req.email_col or "").strip()
        if not email_col:
            raise HTTPException(status_code=400, detail="email_col required")

        temp_path = await stream_input_to_tempfile(supabase, file_path)
        try:
            if file_path.lower().endswith(".xlsx"):
                row_count = count_xlsx_rows(temp_path)
            else:
                row_count = count_csv_rows(temp_path)
            
            # Apply process limit if requested (for partial processing)
            if req.process_limit is not None and req.process_limit > 0:
                row_count = min(row_count, req.process_limit)
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

        job_id = str(uuid.uuid4())

        # Convert service to JSON string if it's a ServiceComponents object
        if isinstance(req.service, ServiceComponents):
            service_str = req.service.model_dump_json()
        elif isinstance(req.service, str):
            # Try to parse as JSON to validate structure, or keep as plain string
            try:
                json.loads(req.service)
                service_str = req.service
            except json.JSONDecodeError:
                # Legacy plain string - keep as is
                service_str = req.service
        else:
            service_str = json.dumps(req.service)

        meta = {
            "file_path": file_path,
            "email_col": email_col,
            "service": service_str,
            "total_rows": row_count,  # Cache row count to avoid re-counting in worker
            "process_limit": req.process_limit,
        }

        lock_name = f"credits_lock:{current_user.user_id}"
        # Tuned lock timeout: 5s total, 2s blocking wait. 
        # 30s was too long for high-throughput.
        lock = redis_conn.lock(lock_name, timeout=5, blocking_timeout=2)
        acquired = lock.acquire(blocking=True)
        if not acquired:
            raise HTTPException(status_code=503, detail="Unable to reserve credits")

        reservation = _reserve_credits_for_job(
            supabase,
            current_user.user_id,
            row_count,
            job_id,
            file_path,
        )

        meta.update(
            {
                "credit_cost": row_count,
                "credits_deducted": True,
                "credits_refunded": False,
                "monthly_deducted": reservation.get("monthly_deducted", 0),
                "addon_deducted": reservation.get("addon_deducted", 0),
            }
        )

        # Generate unique filename
        original_basename = os.path.basename(file_path)
        # Strip timestamp prefix if present (e.g. 1234567890_file.csv -> file.csv)
        clean_name = re.sub(r'^\d+_', '', original_basename)
        
        # Ensure sif_ prefix
        if not clean_name.startswith("sif_"):
            base_candidate = f"sif_{clean_name}"
        else:
            base_candidate = clean_name
            
        # Split name and extension
        name_part, ext_part = os.path.splitext(base_candidate)
        
        # Find existing files with similar names for this user
        # We look for: "name.ext", "name_1.ext", "name_2.ext", etc.
        # Pattern: ^name(_\d+)?\.ext$
        escaped_name = re.escape(name_part)
        escaped_ext = re.escape(ext_part)
        pattern = f"^{escaped_name}(_\\d+)?{escaped_ext}$"
        
        existing_files_res = (
            supabase.table("jobs")
            .select("filename")
            .eq("user_id", current_user.user_id)
            .execute()
        )
        
        existing_filenames = [
            row["filename"] 
            for row in (existing_files_res.data or []) 
            if row.get("filename") and re.match(pattern, row["filename"])
        ]
        
        final_filename = base_candidate
        if base_candidate in existing_filenames:
            counter = 1
            while True:
                candidate = f"{name_part}_{counter}{ext_part}"
                if candidate not in existing_filenames:
                    final_filename = candidate
                    break
                counter += 1

        result = (
            supabase.table("jobs")
            .insert(
                {
                    "id": job_id,
                    "user_id": current_user.user_id,
                    "status": "queued",
                    "filename": final_filename,
                    "rows": row_count,
                    "meta_json": meta,
                }
            )
            .execute()
        )

        if not result.data:
            raise RuntimeError("Failed to insert job")

        job = result.data[0]

        # Publish job notification to Redis for instant worker pickup
        jobs.publish_job_notification(job["id"])

        # Update user's service context
        try:
            # Ensure we have a dict for the service context
            service_data = None
            if isinstance(req.service, ServiceComponents):
                service_data = req.service.model_dump()
            elif isinstance(req.service, str):
                try:
                    service_data = json.loads(req.service)
                except:
                    pass
            
            if service_data:
                supabase.table("profiles").update({
                    "service_context": service_data
                }).eq("id", current_user.user_id).execute()
        except Exception as e:
            print(f"[Job] Failed to update service context: {e}")


        return {"id": job["id"], "status": job["status"], "rows": row_count}

    except HTTPException:
        if reservation is not None:
            _rollback_credit_reservation(
                supabase, current_user.user_id, reservation, job_id
            )
        raise
    except FileStreamingError as exc:
        if reservation is not None:
            _rollback_credit_reservation(
                supabase, current_user.user_id, reservation, job_id
            )
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as e:
        if reservation is not None:
            _rollback_credit_reservation(
                supabase, current_user.user_id, reservation, job_id
            )
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        if lock is not None and acquired:
            try:
                lock.release()
            except Exception:
                pass




@app.get("/jobs/{job_id}")
def get_job(
    job_id: str,
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    supabase = get_supabase()
    user_id = current_user.user_id

    job_res = (
        supabase.table("jobs")
        .select("*")
        .eq("id", job_id)
        .single()
        .execute()
    )
    if not job_res.data:
        raise HTTPException(status_code=404, detail="Job not found")

    job = job_res.data
    if job["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Unauthorized")

    logs_res = (
        supabase.table("job_logs")
        .select("*")
        .eq("job_id", job_id)
        .order("step", desc=True)
        .limit(1)
        .execute()
    )
    last_log = logs_res.data[0] if logs_res.data else None

    if last_log:
        step = last_log["step"]
        total = last_log["total"] or 1
        job["progress"] = int((step / total) * 100)
        job["message"] = last_log.get("message")
    else:
        job["progress"] = 0
        job["message"] = None

    return job


@app.get("/jobs/{job_id}/download")
async def download_result(
    job_id: str,
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    supabase = get_supabase()

    job = (
        supabase.table("jobs")
        .select("user_id,result_path")
        .eq("id", job_id)
        .single()
        .execute()
    )

    if not job.data or not job.data.get("result_path"):
        raise HTTPException(status_code=404, detail="Result not found")

    if job.data.get("user_id") != current_user.user_id:
        raise HTTPException(status_code=403)

    storage_path = job.data["result_path"]
    filename = job.data.get("filename", "result.xlsx")
    
    file = supabase.storage.from_("outputs").download(storage_path)
    return StreamingResponse(
        io.BytesIO(file),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        },
    )


class PreviewEmailsRequest(BaseModel):
    file_path: str
    email_col: str

class GeneratePreviewRequest(BaseModel):
    file_path: str
    email_col: str
    selected_email: str
    service: str  # JSON string of service components


@app.post("/preview/emails")
async def get_preview_emails(
    payload: PreviewEmailsRequest = Body(...),
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    """Extract first 5 emails from the uploaded file for preview selection."""
    try:
        file_path = payload.file_path
        email_col = payload.email_col

        if not file_path or not email_col:
            raise HTTPException(status_code=400, detail="file_path and email_col required")




        # Verify user owns the file
        file_path = assert_user_owns_path(file_path, current_user.user_id)

        supabase_client = get_supabase()
        temp_path = await stream_input_to_tempfile(supabase_client, file_path)

        try:
            emails = []
            if file_path.lower().endswith(".xlsx"):
                from openpyxl import load_workbook
                workbook = load_workbook(temp_path, read_only=True)
                sheet = workbook.active
                rows_iter = sheet.iter_rows(values_only=True)

                # Get header row
                header_row = next(rows_iter, None)
                if not header_row:
                    raise HTTPException(status_code=400, detail="Empty file")

                headers = ["" if cell is None else str(cell) for cell in header_row]
                if email_col not in headers:
                    raise HTTPException(status_code=400, detail=f"Column '{email_col}' not found")

                email_col_idx = headers.index(email_col)

                # Extract emails for preview (limit to first 10)
                for i, row in enumerate(rows_iter):
                    if len(emails) >= 10:
                        break
                    if row and len(row) > email_col_idx:
                        email_value = row[email_col_idx]
                        if email_value:
                            emails.append(str(email_value))

                workbook.close()
            else:
                # CSV file
                import csv
                with open(temp_path, newline="", encoding="utf-8-sig") as handle:
                    reader = csv.DictReader(handle)
                    if email_col not in (reader.fieldnames or []):
                        raise HTTPException(status_code=400, detail=f"Column '{email_col}' not found")

                    for i, row in enumerate(reader):
                        if len(emails) >= 10:
                            break
                        email_value = row.get(email_col, "")
                        if email_value:
                            emails.append(email_value)

            return {"emails": emails}
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

    except HTTPException:
        raise
    except FileStreamingError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as e:
        print(f"[Preview] Unexpected error in get_preview_emails: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/user/settings")
def update_user_settings(
    payload: Dict[str, Any] = Body(...),
    current_user: AuthenticatedUser = Depends(get_current_user)
):
    supabase = get_supabase()
    user_id = current_user.user_id
    
    service_context = payload.get("service_context")
    if service_context is None:
        raise HTTPException(status_code=400, detail="service_context is required")
        
    try:
        supabase.table("profiles").update({
            "service_context": service_context
        }).eq("id", user_id).execute()
        
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/preview/generate")
async def generate_preview(
    payload: GeneratePreviewRequest = Body(...),
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    """Generate a preview email using the full pipeline (SERP + 3 Groq calls) and deduct 1 credit."""
    from .research import perform_research
    from .gpt_helpers import generate_full_email_body
    from .email_cleaning import clean_email_body

    try:
        selected_email = payload.selected_email
        service_context = payload.service

        if not selected_email or "@" not in selected_email:
            raise HTTPException(status_code=400, detail="Valid email required")

        # Verify user owns the file
        file_path = assert_user_owns_path(payload.file_path, current_user.user_id)

        # Deduct 1 credit using the same atomic pattern
        supabase_client = get_supabase()
        user_id = current_user.user_id

        # Atomic credit deduction using two-bucket system (monthly first, then addon)
        max_attempts = 5
        credit_deducted = False
        for attempt in range(max_attempts):
            profile_res = (
                supabase_client.table("profiles")
                .select("credits_remaining, addon_credits")
                .eq("id", user_id)
                .limit(1)
                .execute()
            )
            profile = _parse_profile_response(profile_res) or {}
            try:
                credits_remaining = int(profile.get("credits_remaining") or 0)
                addon_credits = int(profile.get("addon_credits") or 0)
            except (TypeError, ValueError):
                credits_remaining = 0
                addon_credits = 0

            total_credits = credits_remaining + addon_credits

            if total_credits < 1:
                raise HTTPException(
                    status_code=402,
                    detail={
                        "error": "insufficient_credits",
                        "credits_remaining": total_credits,
                        "message": "You need at least 1 credit to generate a preview"
                    }
                )

            # Two-bucket deduction: use monthly credits first, then add-ons
            if credits_remaining >= 1:
                # Use monthly credits
                new_monthly = credits_remaining - 1
                new_addon = addon_credits
                deduction_source = "monthly"
            else:
                # Use addon credits (monthly is 0)
                new_monthly = 0
                new_addon = addon_credits - 1
                deduction_source = "addon"

            update_res = (
                supabase_client.table("profiles")
                .update({
                    "credits_remaining": new_monthly,
                    "addon_credits": new_addon
                })
                .eq("id", user_id)
                .eq("credits_remaining", credits_remaining)
                .eq("addon_credits", addon_credits)
                .execute()
            )

            if getattr(update_res, "error", None):
                raise HTTPException(status_code=500, detail="Unable to deduct credits")

            updated_rows = getattr(update_res, "data", None) or []
            if updated_rows:
                # Record ledger entry with deduction source
                ledger_payload = {
                    "user_id": user_id,
                    "change": -1,
                    "amount": 0.0,
                    "reason": f"preview generation ({deduction_source})",
                    "ts": datetime.utcnow().isoformat(),
                }
                ledger_res = (
                    supabase_client.table("ledger").insert(ledger_payload).execute()
                )
                if getattr(ledger_res, "error", None):
                    # Rollback credit deduction to both buckets
                    supabase_client.table("profiles").update({
                        "credits_remaining": credits_remaining,
                        "addon_credits": addon_credits
                    }).eq("id", user_id).execute()
                    raise HTTPException(
                        status_code=500, detail="Unable to record credit deduction"
                    )

                credit_deducted = True
                print(
                    f"[Preview] Deducted 1 credit from {deduction_source} bucket for user {user_id} "
                    f"(monthly: {credits_remaining}→{new_monthly}, addon: {addon_credits}→{new_addon})"
                )
                break

            time.sleep(0.1)

        if not credit_deducted:
            raise HTTPException(status_code=500, detail="Unable to deduct credits")

        # Run the exact same pipeline as the main job
        print(f"\n[Preview] Generating preview for email: {selected_email}")

        # Step 1: Research (SERP + Groq synthesis)
        research_components = perform_research(selected_email)
        print(f"\n📋 PREVIEW RESEARCH COMPONENTS:")
        print(research_components)

        # Step 2: Email generation (Groq)
        email_body = generate_full_email_body(
            research_components,
            service_context,
        )
        print(f"\n✉️  PREVIEW EMAIL BODY (before cleaning):")
        print(email_body)

        # Step 3: Cleaning (Groq)
        email_body = clean_email_body(email_body)
        print(f"\n🧹 PREVIEW CLEANED EMAIL BODY:")
        print(email_body)

        return {
            "email": selected_email,
            "research_components": research_components,
            "email_body": email_body,
            "credits_remaining": new_monthly + new_addon
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"[Preview] Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/jobs/{job_id}/progress")
def job_progress(
    job_id: str,
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    supabase = get_supabase()
    user_id = current_user.user_id

    job_res = (
        supabase.table("jobs")
        .select("id,user_id,status")
        .eq("id", job_id)
        .single()
        .execute()
    )
    if not job_res.data:
        raise HTTPException(status_code=404, detail="Job not found")

    job = job_res.data
    if job["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Unauthorized")

    logs_res = (
        supabase.table("job_logs")
        .select("*")
        .eq("job_id", job_id)
        .order("step", desc=True)
        .limit(1)
        .execute()
    )
    last_log = logs_res.data[0] if logs_res.data else None

    if last_log:
        step = last_log["step"]
        total = last_log["total"] or 1
        percent = int((step / total) * 100)
        message = last_log.get("message")
    else:
        percent = 0
        message = None

    return {
        "job_id": job_id,
        "status": job["status"],
        "percent": percent,
        "message": message,
    }


# WebSocket endpoint for real-time job progress
@app.websocket("/ws/jobs/{job_id}")
async def websocket_job_progress(websocket: WebSocket, job_id: str):
    """
    WebSocket endpoint for real-time job progress updates.
    Authenticates via query param 'token', then subscribes to Redis pub/sub
    for live updates on the specified job.
    """
    # Check if WebSocket is enabled via feature flag
    enable_websocket = os.getenv("ENABLE_WEBSOCKET", "true").lower() == "true"
    if not enable_websocket:
        await websocket.close(code=1008, reason="WebSocket not enabled")
        return

    # Extract token from query params
    query_params = dict(websocket.query_params)
    token = query_params.get("token")

    if not token:
        await websocket.close(code=1008, reason="Missing authentication token")
        return

    # Validate JWT token
    try:
        secret = os.getenv("SUPABASE_JWT_SECRET")
        if not secret:
            await websocket.close(code=1011, reason="Server configuration error")
            return

        supabase_url = os.getenv("SUPABASE_URL", "").rstrip("/")
        issuer = os.getenv("SUPABASE_JWT_ISS", f"{supabase_url}/auth/v1" if supabase_url else None)
        audience = os.getenv("SUPABASE_JWT_AUD", "authenticated")

        payload = jwt.decode(
            token,
            secret,
            algorithms=["HS256"],
            issuer=issuer,
            audience=audience,
            options={"verify_exp": True},
        )
        user_id = payload.get("sub")
        if not user_id:
            await websocket.close(code=1008, reason="Invalid token")
            return

    except ExpiredSignatureError:
        await websocket.close(code=1008, reason="Token expired")
        return
    except (InvalidTokenError, InvalidAudienceError, InvalidIssuerError) as e:
        await websocket.close(code=1008, reason="Invalid token")
        return
    except Exception as e:
        logging.error(f"WebSocket auth error: {e}")
        await websocket.close(code=1011, reason="Authentication failed")
        return

    # Verify user owns this job
    try:
        supabase_client = get_supabase()
        job_res = (
            supabase_client.table("jobs")
            .select("id,user_id,status")
            .eq("id", job_id)
            .single()
            .execute()
        )
        if not job_res.data:
            await websocket.close(code=1008, reason="Job not found")
            return

        job = job_res.data
        if job["user_id"] != user_id:
            await websocket.close(code=1008, reason="Unauthorized")
            return
    except Exception as e:
        logging.error(f"WebSocket job verification error: {e}")
        await websocket.close(code=1011, reason="Verification failed")
        return

    # Accept the WebSocket connection
    await websocket.accept()

    # Create Redis pub/sub connection
    pubsub = redis_conn.pubsub()
    channel = f"job_progress:{job_id}"

    try:
        # Subscribe to job progress channel
        pubsub.subscribe(channel)

        # Send initial status
        initial_status = {
            "job_id": job_id,
            "status": job["status"],
            "percent": 0,
            "message": "Connected to job progress stream",
        }
        await websocket.send_json(initial_status)

        # Listen for Redis pub/sub messages in a background task
        async def listen_redis():
            """Listen to Redis pub/sub and forward to WebSocket"""
            loop = asyncio.get_event_loop()

            while True:
                # Run blocking Redis get_message in executor
                message = await loop.run_in_executor(
                    None,
                    pubsub.get_message,
                    True,  # ignore_subscribe_messages
                    1.0    # timeout in seconds
                )

                if message and message['type'] == 'message':
                    try:
                        # Parse the Redis message data
                        data = json.loads(message['data'])
                        await websocket.send_json(data)

                        # If job is complete, close connection
                        if data.get('status') in ['succeeded', 'failed']:
                            await asyncio.sleep(1)  # Give client time to receive final message
                            break
                    except json.JSONDecodeError:
                        logging.error(f"Failed to parse Redis message: {message['data']}")
                    except Exception as e:
                        logging.error(f"Error sending WebSocket message: {e}")
                        break

                await asyncio.sleep(0.1)  # Small delay to prevent tight loop

        # Run the Redis listener
        await listen_redis()

    except WebSocketDisconnect:
        logging.info(f"WebSocket disconnected for job {job_id}")
    except Exception as e:
        logging.error(f"WebSocket error for job {job_id}: {e}")
    finally:
        # Clean up
        try:
            pubsub.unsubscribe(channel)
            pubsub.close()
        except:
            pass

        try:
            await websocket.close()
        except:
            pass


# ✅ Stripe Checkout Session (subscription OR add-on)
class CheckoutRequest(BaseModel):
    plan: str
    addon: bool = False
    quantity: int = 1

# ========================
# /create_checkout_session
# ========================
@app.post("/create_checkout_session")
async def create_checkout_session(
    data: CheckoutRequest,
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    try:
        print("========== CREATE CHECKOUT SESSION ==========")
        print(f"[DEBUG] Incoming data: {data}")
        user_id = current_user.user_id
        user_email = current_user.claims.get("email")

        # Rate Limiting: 5 requests per minute per user
        rate_limit_key = f"rate_limit:checkout:{user_id}"
        current_requests = redis_conn.incr(rate_limit_key)
        if current_requests == 1:
            redis_conn.expire(rate_limit_key, 60)
        
        if current_requests > 5:
            print(f"[RATE_LIMIT] User {user_id} exceeded checkout limit")
            raise HTTPException(status_code=429, detail="Too many checkout requests. Please try again in a minute.")

        if not STRIPE_SECRET:
            raise HTTPException(status_code=500, detail="Stripe is not configured")

        # Fetch current profile to check existing subscription
        supabase_client = get_supabase()
        profile_res = (
            supabase_client.table("profiles")
            .select("plan_type, subscription_status")
            .eq("id", user_id)
            .single()
            .execute()
        )
        
        current_plan = "free"
        subscription_status = None
        
        if profile_res.data:
            current_plan = profile_res.data.get("plan_type", "free")
            subscription_status = profile_res.data.get("subscription_status")

        # Normalize plans for comparison (remove _monthly/_annual suffixes)
        def normalize_plan(p: str) -> str:
            return p.lower().replace("_annual", "").replace("_monthly", "")

        requested_plan_norm = normalize_plan(data.plan)
        current_plan_norm = normalize_plan(current_plan)

        print(f"[DEBUG] Validation Check - User: {user_id}")
        print(f"[DEBUG] Current Plan (DB): {current_plan} -> Norm: {current_plan_norm}")
        print(f"[DEBUG] Requested Plan: {data.plan} -> Norm: {requested_plan_norm}")
        print(f"[DEBUG] Subscription Status: {subscription_status}")
        print(f"[DEBUG] Is Addon: {data.addon}")

        # Block purchasing the same plan if subscription is active
        if (
            not data.addon 
            and subscription_status == "active" 
            and requested_plan_norm == current_plan_norm
            and current_plan != "free"
        ):
            print(f"[BLOCK] User {user_id} attempted to buy current plan: {data.plan} (Current: {current_plan})")
            raise HTTPException(
                status_code=400, 
                detail=f"You are already subscribed to the {requested_plan_norm.capitalize()} plan."
            )

        stripe_customer_id = ensure_stripe_customer_id(user_id, email=user_email)
        if not stripe_customer_id:
            raise HTTPException(status_code=500, detail="Unable to create Stripe customer")

        line_items = []
        price_id = None
        mode = "subscription"

        if data.addon:
            addon_price_id = ADDON_PRICE_MAP.get(data.plan)
            if not addon_price_id:
                print(f"[ERROR] Invalid addon plan: {data.plan}")
                return {"error": "Invalid addon plan"}

            if data.quantity < 1:
                 raise HTTPException(status_code=400, detail="Minimum addon quantity is 1")

            price_id = addon_price_id
            line_items = [
                {
                    "price": addon_price_id,
                    "quantity": data.quantity,
                }
            ]
            mode = "payment"
        else:
            # Check if annual plan (ends with "_annual")
            if data.plan.endswith("_annual"):
                price_id = PLAN_PRICE_MAP_ANNUAL.get(data.plan)
                billing_type = "annual"
            else:
                price_id = PLAN_PRICE_MAP.get(data.plan)
                billing_type = "monthly"

            if not price_id:
                print(f"[ERROR] Invalid plan: {data.plan}")
                return {"error": "Invalid plan"}

            line_items = [
                {
                    "price": price_id,
                    "quantity": 1,
                }
            ]
            mode = "subscription"
            print(f"[INFO] Creating {billing_type} subscription for plan: {data.plan}")

        print(
            f"[INFO] Using price_id={price_id}, mode={mode}, user_id={user_id}, customer={stripe_customer_id}"
        )

        session = stripe.checkout.Session.create(
            customer=stripe_customer_id,
            payment_method_types=["card"],
            mode=mode,
            line_items=line_items,
            success_url=SUCCESS_URL,
            cancel_url=CANCEL_URL,
            metadata={
                "user_id": user_id,
                "plan": data.plan,
                "addon": str(data.addon).lower(),
                "quantity": str(data.quantity),
            },
            subscription_data=None
            if data.addon
            else {
                "metadata": {
                    "user_id": user_id,
                    "plan": data.plan,
                }
            },
        )

        print(f"[INFO] Created Checkout Session: {session['id']}")
        print("========== END CREATE CHECKOUT SESSION ==========")

        return {"id": session.id}

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] create_checkout_session failed: {e}")
        return {"error": str(e)}



@app.post("/stripe/sync")
async def trigger_stripe_sync(
    current_user: AuthenticatedUser = Depends(get_current_user),
):
    if not STRIPE_SECRET:
        raise HTTPException(status_code=500, detail="Stripe is not configured")

    user_id = current_user.user_id
    stripe_customer_id = ensure_stripe_customer_id(
        user_id, email=current_user.claims.get("email")
    )

    if not stripe_customer_id:
        raise HTTPException(status_code=400, detail="Stripe customer missing")

    sync_result = sync_stripe_customer(stripe_customer_id, user_id=user_id)
    return {
        "status": "ok",
        "customerId": stripe_customer_id,
        "subscription": sync_result.get("subscription"),
    }



# Init Supabase (must be service role key)
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET")

def log_db(label, res):
    print(f"[DB:{label}] {res}")

def _is_event_processed(event_id: str) -> bool:
    """Check if a Stripe event has already been processed (idempotency check)"""
    try:
        result = (
            supabase.table("processed_stripe_events")
            .select("event_id")
            .eq("event_id", event_id)
            .limit(1)
            .execute()
        )
        return result.data is not None and len(result.data) > 0
    except Exception as e:
        print(f"[IDEMPOTENCY] Error checking event {event_id}: {e}")
        # If table doesn't exist or error, allow processing (fail open)
        return False

def _mark_event_processed(event_id: str, event_type: str) -> None:
    """Mark a Stripe event as processed"""
    try:
        supabase.table("processed_stripe_events").insert({
            "event_id": event_id,
            "event_type": event_type,
            "processed_at": datetime.utcnow().isoformat(),
        }).execute()
    except Exception as e:
        print(f"[IDEMPOTENCY] Error marking event {event_id}: {e}")

def _verify_payment_status(obj) -> bool:
    """Verify that payment was actually successful"""
    # For checkout sessions
    if obj.get("object") == "checkout.session":
        payment_status = obj.get("payment_status")
        return payment_status == "paid"

    # For invoices
    if obj.get("object") == "invoice":
        status = obj.get("status")
        paid = obj.get("paid")
        return status == "paid" and paid == True

    return False


def _revoke_all_credits_for_fraud(
    user_id: str,
    reason: str,
    event_type: str,
    event_id: str,
    amount: float = 0.0,
    cancel_subscription: bool = True,
) -> dict:
    """
    Revoke ALL credits (monthly + addon) for a user due to fraud/refund/dispute.

    This is the nuclear option - used when a user bypasses the app and goes directly
    to Stripe for refunds or files disputes. Unlike normal cancellation (where users
    keep credits until cycle end), this immediately revokes everything.

    Args:
        user_id: The user to revoke credits from
        reason: Human-readable reason for the revocation
        event_type: Stripe event type that triggered this
        event_id: Stripe event ID for audit trail
        amount: Dollar amount involved (for logging)
        cancel_subscription: Whether to also cancel their Stripe subscription

    Returns:
        dict with revocation details
    """
    try:
        # Get current credits before revocation
        profile_res = (
            supabase.table("profiles")
            .select("credits_remaining, addon_credits, plan_type, stripe_subscription_id, stripe_customer_id")
            .eq("id", user_id)
            .single()
            .execute()
        )

        if not profile_res.data:
            print(f"[FRAUD_REVOKE] Profile not found for user {user_id}")
            return {"status": "error", "message": "profile_not_found"}

        profile = profile_res.data
        old_monthly = profile.get("credits_remaining", 0)
        old_addon = profile.get("addon_credits", 0)
        old_plan = profile.get("plan_type", "free")
        stripe_sub_id = profile.get("stripe_subscription_id")
        total_revoked = old_monthly + old_addon

        # Revoke ALL credits and mark account
        update_payload = {
            "credits_remaining": 0,
            "addon_credits": 0,
            "subscription_status": "revoked",
            "plan_type": "free",
            "pending_downgrade": None,
            "pending_downgrade_effective_date": None,
            "fraud_flag": event_type,  # Mark what triggered the revocation
            "fraud_flag_date": datetime.utcnow().isoformat(),
        }

        res_update = (
            supabase.table("profiles")
            .update(update_payload)
            .eq("id", user_id)
            .execute()
        )

        # Log to ledger with full details
        ledger_entry = {
            "user_id": user_id,
            "change": -total_revoked,
            "amount": -abs(amount) if amount else 0,
            "reason": f"FRAUD_REVOKE: {reason} | event={event_type} | monthly=-{old_monthly}, addon=-{old_addon}, plan={old_plan}",
            "ts": datetime.utcnow().isoformat(),
        }
        supabase.table("ledger").insert(ledger_entry).execute()

        # Also invalidate any addon_credit_purchases
        try:
            supabase.table("addon_credit_purchases").update({
                "credits_remaining": 0,
                "revoked_at": datetime.utcnow().isoformat(),
                "revoked_reason": f"{event_type}: {reason}",
            }).eq("user_id", user_id).gt("credits_remaining", 0).execute()
        except Exception as e:
            print(f"[FRAUD_REVOKE] Warning: Could not invalidate addon purchases: {e}")

        # Cancel Stripe subscription if requested and exists
        subscription_canceled = False
        if cancel_subscription and stripe_sub_id:
            try:
                stripe.Subscription.cancel(stripe_sub_id)
                subscription_canceled = True
                print(f"[FRAUD_REVOKE] Canceled Stripe subscription {stripe_sub_id}")
            except stripe.error.InvalidRequestError as e:
                # Subscription might already be canceled
                print(f"[FRAUD_REVOKE] Subscription already canceled or invalid: {e}")
                subscription_canceled = True  # Consider it done
            except Exception as e:
                print(f"[FRAUD_REVOKE] Warning: Could not cancel subscription: {e}")

        result = {
            "status": "revoked",
            "user_id": user_id,
            "monthly_revoked": old_monthly,
            "addon_revoked": old_addon,
            "total_revoked": total_revoked,
            "old_plan": old_plan,
            "subscription_canceled": subscription_canceled,
            "event_type": event_type,
            "event_id": event_id,
        }

        print(
            f"[FRAUD_REVOKE] ⚠️  REVOKED ALL CREDITS for user {user_id}: "
            f"monthly={old_monthly}, addon={old_addon}, total={total_revoked}, "
            f"plan={old_plan}→free, reason={reason}"
        )

        return result

    except Exception as e:
        print(f"[FRAUD_REVOKE] ERROR revoking credits for user {user_id}: {e}")
        return {"status": "error", "message": str(e)}

@app.post("/stripe-webhook")
async def stripe_webhook(request: Request, stripe_signature: str = Header(None)):
    payload = await request.body()
    print("========== STRIPE WEBHOOK ==========")

    try:
        event = stripe.Webhook.construct_event(
            payload, stripe_signature, STRIPE_WEBHOOK_SECRET
        )
    except Exception as e:
        print(f"[ERROR] Webhook verification failed: {e}")
        raise HTTPException(status_code=400, detail=f"Webhook error: {e}")

    event_type = event["type"]
    obj = event["data"]["object"]
    event_id = event["id"]
    print(f"[EVENT] {event_type} id={event_id}")

    # Idempotency check - prevent processing the same event twice
    if _is_event_processed(event_id):
        print(f"[IDEMPOTENCY] Event {event_id} already processed, skipping")
        return {"status": "duplicate", "event_id": event_id}

    def log_db(label, res):
        print(f"[DB:{label}] {res}")

    metadata = obj.get("metadata", {}) if hasattr(obj, "get") else {}
    if hasattr(metadata, "to_dict"):
        metadata = metadata.to_dict()
    customer_id = obj.get("customer") if hasattr(obj, "get") else None
    linked_user_id: Optional[str] = None

    # Payment verification - ensure payment was actually successful before granting credits
    payment_verified = _verify_payment_status(obj)
    if event_type in ["checkout.session.completed", "invoice.paid"] and not payment_verified:
        print(f"[WARNING] Payment not verified for event {event_id}, skipping credit allocation")
        _mark_event_processed(event_id, event_type)
        return {"status": "payment_not_verified", "event_id": event_id}

    # Handle payment failures
    if event_type == "invoice.payment_failed":
        subscription_id = obj.get("subscription")
        sub_metadata = obj.get("metadata", {}) # Invoice metadata might not have it, check subscription later
        
        if subscription_id:
             try:
                subscription = stripe.Subscription.retrieve(subscription_id)
                user_id = subscription.get("metadata", {}).get("user_id")
                
                if not user_id and customer_id:
                     user_id = _find_user_by_customer(customer_id)
                
                if user_id:
                    print(f"[PAYMENT_FAILED] Marking user {user_id} as past_due")
                    supabase.table("profiles").update({
                        "subscription_status": "past_due"
                    }).eq("id", user_id).execute()
                    
                    # Log to ledger for audit trail
                    supabase.table("ledger").insert({
                        "user_id": user_id,
                        "change": 0,
                        "amount": 0,
                        "reason": f"payment_failed - subscription {subscription_id}",
                        "ts": datetime.utcnow().isoformat(),
                    }).execute()
             except Exception as e:
                 print(f"[PAYMENT_FAILED] Error handling payment failure: {e}")
        
        # Mark event as processed
        _mark_event_processed(event_id, event_type)
        return {"status": "payment_failure_handled", "event_id": event_id}

    # ============================================================
    # FRAUD PROTECTION: Handle refunds, disputes, and chargebacks
    # Policy: Direct Stripe actions (bypassing app) = IMMEDIATE revocation of ALL credits
    # This is different from in-app cancellation where users keep credits until cycle end
    # ============================================================

    # Handle charge refunds (user requested refund directly from Stripe/bank)
    if event_type == "charge.refunded":
        print(f"[REFUND] ⚠️  charge.refunded detected - REVOKING ALL CREDITS")

        try:
            if customer_id:
                user_id = _find_user_by_customer(customer_id)
                if user_id:
                    linked_user_id = user_id
                    amount_refunded = obj.get("amount_refunded", 0) / 100.0
                    charge_id = obj.get("id", "unknown")

                    revoke_result = _revoke_all_credits_for_fraud(
                        user_id=user_id,
                        reason=f"Charge refunded (${amount_refunded:.2f}) - charge_id={charge_id}",
                        event_type=event_type,
                        event_id=event_id,
                        amount=amount_refunded,
                        cancel_subscription=True,
                    )

                    _mark_event_processed(event_id, event_type)
                    return {"status": "credits_revoked", "event_id": event_id, "details": revoke_result}
                else:
                    print(f"[REFUND] Could not find user for customer {customer_id}")
            else:
                print(f"[REFUND] No customer_id in refund event")
        except Exception as e:
            print(f"[REFUND] Error handling refund: {e}")

        _mark_event_processed(event_id, event_type)
        return {"status": "refund_handled", "event_id": event_id}

    # Handle disputes/chargebacks (user disputed with their bank)
    if event_type == "charge.dispute.created":
        print(f"[DISPUTE] ⚠️  charge.dispute.created detected - REVOKING ALL CREDITS")

        try:
            if customer_id:
                user_id = _find_user_by_customer(customer_id)
                if user_id:
                    linked_user_id = user_id
                    dispute_amount = obj.get("amount", 0) / 100.0
                    dispute_reason = obj.get("reason", "unknown")
                    dispute_id = obj.get("id", "unknown")

                    revoke_result = _revoke_all_credits_for_fraud(
                        user_id=user_id,
                        reason=f"Dispute/Chargeback filed (${dispute_amount:.2f}) - reason={dispute_reason}, dispute_id={dispute_id}",
                        event_type=event_type,
                        event_id=event_id,
                        amount=dispute_amount,
                        cancel_subscription=True,
                    )

                    _mark_event_processed(event_id, event_type)
                    return {"status": "credits_revoked_dispute", "event_id": event_id, "details": revoke_result}
                else:
                    print(f"[DISPUTE] Could not find user for customer {customer_id}")
            else:
                print(f"[DISPUTE] No customer_id in dispute event")
        except Exception as e:
            print(f"[DISPUTE] Error handling dispute: {e}")

        _mark_event_processed(event_id, event_type)
        return {"status": "dispute_handled", "event_id": event_id}

    # Handle dispute resolution (if we lose, ensure credits stay revoked)
    if event_type == "charge.dispute.closed":
        print(f"[DISPUTE_CLOSED] charge.dispute.closed detected")

        try:
            dispute_status = obj.get("status", "unknown")  # won, lost, needs_response, etc.

            if customer_id:
                user_id = _find_user_by_customer(customer_id)
                if user_id:
                    linked_user_id = user_id
                    dispute_amount = obj.get("amount", 0) / 100.0
                    dispute_id = obj.get("id", "unknown")

                    if dispute_status == "lost":
                        # Ensure credits are revoked (they should already be from dispute.created)
                        print(f"[DISPUTE_CLOSED] Dispute LOST - ensuring credits revoked for user {user_id}")
                        revoke_result = _revoke_all_credits_for_fraud(
                            user_id=user_id,
                            reason=f"Dispute LOST (${dispute_amount:.2f}) - dispute_id={dispute_id}",
                            event_type=event_type,
                            event_id=event_id,
                            amount=dispute_amount,
                            cancel_subscription=True,
                        )
                    elif dispute_status == "won":
                        # We won the dispute - log it but don't restore credits automatically
                        # Manual review should be required to restore access
                        print(f"[DISPUTE_CLOSED] Dispute WON for user {user_id} - manual review required to restore")
                        supabase.table("ledger").insert({
                            "user_id": user_id,
                            "change": 0,
                            "amount": dispute_amount,
                            "reason": f"Dispute WON - dispute_id={dispute_id} - MANUAL REVIEW REQUIRED to restore credits",
                            "ts": datetime.utcnow().isoformat(),
                        }).execute()
                    else:
                        print(f"[DISPUTE_CLOSED] Dispute status={dispute_status} for user {user_id}")
                        supabase.table("ledger").insert({
                            "user_id": user_id,
                            "change": 0,
                            "amount": 0,
                            "reason": f"Dispute closed with status={dispute_status} - dispute_id={dispute_id}",
                            "ts": datetime.utcnow().isoformat(),
                        }).execute()
        except Exception as e:
            print(f"[DISPUTE_CLOSED] Error handling dispute closure: {e}")

        _mark_event_processed(event_id, event_type)
        return {"status": "dispute_closed_handled", "event_id": event_id}

    # Handle early fraud warnings (Stripe Radar detected potential fraud)
    if event_type == "radar.early_fraud_warning.created":
        print(f"[FRAUD_WARNING] ⚠️  Early fraud warning detected - REVOKING ALL CREDITS")

        try:
            charge_id = obj.get("charge")
            if charge_id:
                # Get the charge to find the customer
                try:
                    charge = stripe.Charge.retrieve(charge_id)
                    customer_id = charge.get("customer")
                except Exception as e:
                    print(f"[FRAUD_WARNING] Could not retrieve charge {charge_id}: {e}")
                    customer_id = None

            if customer_id:
                user_id = _find_user_by_customer(customer_id)
                if user_id:
                    linked_user_id = user_id
                    fraud_type = obj.get("fraud_type", "unknown")

                    revoke_result = _revoke_all_credits_for_fraud(
                        user_id=user_id,
                        reason=f"Early fraud warning - type={fraud_type}, charge={charge_id}",
                        event_type=event_type,
                        event_id=event_id,
                        amount=0,
                        cancel_subscription=True,
                    )

                    _mark_event_processed(event_id, event_type)
                    return {"status": "credits_revoked_fraud_warning", "event_id": event_id, "details": revoke_result}
        except Exception as e:
            print(f"[FRAUD_WARNING] Error handling fraud warning: {e}")

        _mark_event_processed(event_id, event_type)
        return {"status": "fraud_warning_handled", "event_id": event_id}

    if event_type == "checkout.session.completed":
        user_id = metadata.get("user_id")
        plan = metadata.get("plan")
        linked_user_id = user_id

        # CRITICAL: Validate user_id exists
        if not user_id:
            print(f"[ERROR] No user_id in checkout.session.completed metadata for event {event_id}")
            _mark_event_processed(event_id, event_type)
            return {"status": "error", "message": "missing_user_id"}

        if customer_id and user_id:
            _update_profile(user_id, {"stripe_customer_id": customer_id})

        if metadata.get("addon") == "true":
            qty = int(metadata.get("quantity", 1))
            credits = qty * 1000

            # Get user's current plan to determine expiry
            profile_res = supabase.table("profiles").select("plan_type, addon_credits").eq("id", user_id).single().execute()
            if not profile_res.data:
                print(f"[ERROR] Profile not found for user {user_id}")
                _mark_event_processed(event_id, event_type)
                return {"status": "error", "message": "profile_not_found"}

            plan_type = profile_res.data.get("plan_type", "free")
            current_addon = profile_res.data.get("addon_credits", 0)

            try:
                # Calculate expiration based on plan (6 months for Starter, 12 months for Growth/Pro)
                purchased_at = datetime.utcnow()
                expiry_months = ADDON_EXPIRATION_MONTHS.get(plan_type, 12)
                expires_at = purchased_at + timedelta(days=30 * expiry_months)

                # Insert into addon_credit_purchases table
                addon_purchase = supabase.table("addon_credit_purchases").insert({
                    "user_id": user_id,
                    "credits_purchased": credits,
                    "credits_remaining": credits,
                    "price_paid": (obj.get("amount_total") or 0) / 100.0,
                    "purchased_under_plan": plan_type,
                    "purchased_at": purchased_at.isoformat(),
                    "expires_at": expires_at.isoformat(),
                    "stripe_checkout_session_id": obj.get("id"),
                    "stripe_payment_intent": obj.get("payment_intent"),
                }).execute()
                log_db("insert_addon_purchase", addon_purchase)

                # Also update legacy addon_credits column (dual-write for backward compatibility)
                new_addon_total = (current_addon or 0) + credits
                res_update = (
                    supabase.table("profiles")
                    .update({"addon_credits": new_addon_total})
                    .eq("id", user_id)
                    .execute()
                )
                log_db("update_addon_credits_legacy", res_update)

                # Ledger entry with expiry information
                res_ledger = (
                    supabase.table("ledger")
                    .insert({
                        "user_id": user_id,
                        "change": credits,
                        "amount": (obj.get("amount_total") or 0) / 100.0,
                        "reason": f"addon purchase x{qty} (expires {expires_at.strftime('%Y-%m-%d')})",
                        "ts": datetime.utcnow().isoformat(),
                    })
                    .execute()
                )
                log_db("insert_ledger_addon", res_ledger)

                print(
                    f"[ADDON] user={user_id}, qty={qty}, credits={credits}, "
                    f"plan={plan_type}, expires={expires_at.isoformat()}, customer={customer_id}"
                )
            except Exception as exc:  # pragma: no cover - supabase availability
                print(f"[Stripe] Failed to record add-on purchase: {exc}")
        else:
            # Normalize plan name (strip "_annual" suffix for validation)
            base_plan = plan.replace("_annual", "") if plan.endswith("_annual") else plan
            is_annual = plan.endswith("_annual")
            billing_frequency = "annual" if is_annual else "monthly"

            # Validate plan exists
            if base_plan not in CREDITS_MAP:
                print(f"[ERROR] Invalid plan '{plan}' (base: '{base_plan}') in checkout.session.completed for event {event_id}")
                _mark_event_processed(event_id, event_type)
                return {"status": "error", "message": "invalid_plan"}

            subscription_id = obj.get("subscription")
            renewal_date = None
            if subscription_id:
                subscription = stripe.Subscription.retrieve(subscription_id)
                renewal_date = subscription.get("current_period_end")

            # Annual plans get 12x credits upfront, monthly plans get regular allocation
            if is_annual:
                credits = ANNUAL_CREDITS_MAP.get(base_plan, 0)
                credit_description = f"{credits:,} credits upfront (annual)"
            else:
                credits = CREDITS_MAP.get(base_plan, 0)
                credit_description = f"{credits:,} credits/month (monthly)"
            try:
                res_update = (
                    supabase.table("profiles")
                    .update(
                        {
                            "plan_type": base_plan,  # Store base plan name (e.g., "starter", not "starter_annual")
                            "billing_frequency": billing_frequency,  # Store "monthly" or "annual"
                            "subscription_status": "active",
                            "renewal_date": renewal_date,
                            "credits_remaining": credits,
                        }
                    )
                    .eq("id", user_id)
                    .execute()
                )
                log_db("update_profile_plan", res_update)

                res_ledger = (
                    supabase.table("ledger")
                    .insert(
                        {
                            "user_id": user_id,
                            "change": credits,
                            "amount": (obj.get("amount_total") or 0) / 100.0,
                            "reason": f"plan purchase - {base_plan} ({billing_frequency})",
                            "ts": datetime.utcnow().isoformat(),
                        }
                    )
                    .execute()
                )
                log_db("insert_ledger_plan", res_ledger)
            except Exception as exc:  # pragma: no cover - supabase availability
                print(f"[Stripe] Failed to update plan purchase: {exc}")

            print(
                f"[PLAN] user={user_id}, plan={base_plan}, frequency={billing_frequency}, "
                f"credits={credits} ({credit_description}), renewal={renewal_date}"
            )

    # Handle monthly renewals (invoice.paid event)
    elif event_type == "invoice.paid":
        # Get subscription details from the invoice
        subscription_id = obj.get("subscription")
        billing_reason = obj.get("billing_reason")
        
        print(f"[INVOICE] billing_reason={billing_reason}, subscription_id={subscription_id}")

        # Skip initial subscription invoice (already handled by checkout.session.completed)
        if billing_reason == "subscription_create":
            print(f"[INVOICE] Skipping subscription_create invoice (already handled by checkout)")
        elif billing_reason == "subscription_update":
            # This is an upgrade/downgrade invoice - check if it was API-initiated
            # If API-initiated, credits were already handled there
            print(f"[INVOICE] Skipping subscription_update invoice (upgrade credits handled by API)")
        elif not subscription_id:
            print(f"[INVOICE] No subscription_id in invoice, skipping")
        else:
            try:
                # Retrieve subscription to get metadata with user_id
                subscription = stripe.Subscription.retrieve(subscription_id)
                sub_metadata = subscription.get("metadata", {})
                user_id = sub_metadata.get("user_id")

                if not user_id and customer_id:
                    # Fallback: find user by customer_id
                    user_id = _find_user_by_customer(customer_id)

                if not user_id:
                    print(f"[INVOICE] Could not find user_id for subscription {subscription_id}")
                else:
                    linked_user_id = user_id

                    # Get the user's current plan AND pending_downgrade
                    profile = (
                        supabase.table("profiles")
                        .select("plan_type, subscription_status, billing_frequency, credits_remaining, pending_downgrade")
                        .eq("id", user_id)
                        .single()
                        .execute()
                    )

                    if not profile.data:
                        print(f"[INVOICE] Profile not found for user {user_id}")
                    else:
                        print(f"[INVOICE] Found profile for {user_id}: {profile.data}")
                        plan_type = profile.data.get("plan_type", "free")
                        current_credits = profile.data.get("credits_remaining", 0)
                        subscription_status = profile.data.get("subscription_status")
                        pending_downgrade = profile.data.get("pending_downgrade")
                        
                        print(f"[INVOICE] Plan: {plan_type}, Status: {subscription_status}, Credits: {current_credits}, Pending: {pending_downgrade}")

                        # If there's a pending downgrade, apply it NOW
                        if pending_downgrade and pending_downgrade in PLAN_PRICE_MAP:
                            print(f"[INVOICE] Applying pending downgrade: {plan_type} → {pending_downgrade}")
                            try:
                                # Get subscription items to modify
                                items = stripe.SubscriptionItem.list(subscription=subscription_id, limit=1)
                                if items.data:
                                    item_id = items.data[0].id
                                    new_price_id = PLAN_PRICE_MAP[pending_downgrade]
                                    
                                    # NOW modify the Stripe subscription
                                    stripe.Subscription.modify(
                                        subscription_id,
                                        items=[{"id": item_id, "price": new_price_id}],
                                        proration_behavior="none",
                                    )
                                    print(f"[INVOICE] Stripe subscription modified to {pending_downgrade}")
                                    
                                    # Update profile with new plan and clear pending
                                    plan_type = pending_downgrade  # Use the new plan for credit calculation
                                    
                                supabase.table("profiles").update({
                                    "pending_downgrade": None,
                                    "pending_downgrade_effective_date": None,
                                }).eq("id", user_id).execute()
                                
                            except Exception as e:
                                print(f"[INVOICE] Error applying downgrade: {e}")

                        # Use stored billing_frequency instead of querying Stripe
                        billing_frequency = profile.data.get("billing_frequency")
                        is_annual_subscription = billing_frequency == "annual"

                        if is_annual_subscription:
                            print(
                                f"[INVOICE] Annual subscription detected for user {user_id}. "
                                f"Skipping credit reset (credits allocated upfront at purchase)."
                            )
                            # Annual subscriptions don't get monthly credit resets
                            # They received all 12 months of credits upfront at purchase time
                        else:
                            # For subscription_cycle renewals or plan changes
                            # First, try to get the plan from subscription items (most reliable)
                            items = subscription.get("items", {}).get("data", [])
                            actual_plan = plan_type
                            if items:
                                price_id = items[0].get("price", {}).get("id")
                                detected_plan = PRICE_TO_PLAN.get(price_id)
                                if detected_plan:
                                    actual_plan = detected_plan
                                    print(f"[INVOICE] Detected plan from subscription: {actual_plan}")
                            
                            old_plan_type = plan_type
                            is_plan_change = False
                            
                            # Check if this is a plan change
                            if billing_reason == "subscription_update" or actual_plan != plan_type:
                                if actual_plan != plan_type:
                                    print(f"[PLAN_CHANGE] {plan_type} → {actual_plan}")
                                    is_plan_change = True
                                    plan_type = actual_plan

                            # Handle renewals for any paid plan
                            # Also handle case where subscription_status might be missing (Dashboard-created subs)
                            if plan_type in CREDITS_MAP and plan_type != "free":
                                monthly_credits = CREDITS_MAP.get(plan_type, 0)
                                old_monthly_credits = CREDITS_MAP.get(old_plan_type, 0)

                                # Handle different scenarios
                                should_reset_credits = True
                                ledger_reason = f"monthly renewal - {plan_type}"

                                if is_plan_change:
                                    # Check if upgrade or downgrade
                                    if monthly_credits > old_monthly_credits:
                                        # UPGRADE: Reset to new plan credits
                                        should_reset_credits = True
                                        ledger_reason = f"Plan upgrade - {old_plan_type} to {plan_type}"
                                        print(
                                            f"[INVOICE_UPGRADE] Applying upgrade credits for "
                                            f"{old_plan_type}→{plan_type}"
                                        )
                                    else:
                                        # DOWNGRADE: Reset to new lower plan amount
                                        ledger_reason = f"Plan downgrade - {old_plan_type} to {plan_type}"
                                        print(
                                            f"[INVOICE_DOWNGRADE] Resetting credits for downgrade "
                                            f"{old_plan_type}→{plan_type}: {current_credits}→{monthly_credits}"
                                        )
                                elif billing_reason == "subscription_cycle":
                                    # Regular monthly renewal
                                    print(f"[INVOICE] Processing subscription_cycle renewal for {plan_type}")
                                    ledger_reason = f"monthly renewal - {plan_type}"

                                if should_reset_credits:
                                    # Check if rollover is enabled for this plan
                                    rollover_config = ROLLOVER_RULES.get(plan_type, {"enabled": False, "max_multiplier": 1})

                                    if rollover_config["enabled"] and billing_reason == "subscription_cycle":
                                        # ROLLOVER ENABLED (Growth/Pro): Add to existing, cap at 2x
                                        new_total = current_credits + monthly_credits
                                        max_allowed = monthly_credits * rollover_config["max_multiplier"]
                                        final_credits = min(new_total, max_allowed)

                                        # Calculate how much was lost to cap (if any)
                                        rollover_lost = max(0, new_total - max_allowed)

                                        ledger_reason = f"monthly renewal - {plan_type} (rollover: {current_credits} + {monthly_credits} = {final_credits})"

                                        if rollover_lost > 0:
                                            print(f"[ROLLOVER_CAP] User {user_id} lost {rollover_lost} credits (capped at {max_allowed})")
                                    else:
                                        # NO ROLLOVER (Starter) or plan change: Hard reset
                                        final_credits = monthly_credits
                                        if billing_reason == "subscription_cycle":
                                            ledger_reason = f"monthly renewal - {plan_type} (no rollover, reset to {monthly_credits})"

                                    # Update credits and ensure subscription is marked active
                                    res_update = (
                                        supabase.table("profiles")
                                        .update({
                                            "credits_remaining": final_credits,
                                            "plan_type": plan_type,
                                            "subscription_status": "active",  # Ensure status is active on successful renewal
                                        })
                                        .eq("id", user_id)
                                        .execute()
                                    )
                                    log_db("reset_monthly_credits", res_update)

                                    # Record in ledger
                                    invoice_amount = (obj.get("amount_paid") or 0) / 100.0
                                    res_ledger = (
                                        supabase.table("ledger")
                                        .insert({
                                            "user_id": user_id,
                                            "change": monthly_credits,  # Show the allocation amount
                                            "amount": invoice_amount,
                                            "reason": ledger_reason,
                                            "ts": datetime.utcnow().isoformat(),
                                        })
                                        .execute()
                                    )
                                    log_db("insert_ledger_renewal", res_ledger)

                                    print(f"[RENEWAL] user={user_id}, plan={plan_type}, "
                                          f"old={current_credits}, new={final_credits}, rollover_enabled={rollover_config['enabled']}")
                            else:
                                print(
                                    f"[INVOICE] Skipping credit reset for user {user_id}: "
                                    f"plan={plan_type} (free or unknown)"
                                )
            except Exception as exc:
                print(f"[INVOICE] Error processing monthly renewal: {exc}")

    # Handle subscription updates (upgrades/downgrades)
    elif event_type == "customer.subscription.updated":
        subscription = obj
        subscription_id = subscription.get("id")
        sub_metadata = subscription.get("metadata", {})
        user_id = sub_metadata.get("user_id")

        if not user_id and customer_id:
            user_id = _find_user_by_customer(customer_id)

        if user_id:
            linked_user_id = user_id
            try:
                # Get current profile to compare plans
                profile = (
                    supabase.table("profiles")
                    .select("plan_type, credits_remaining, subscription_status")
                    .eq("id", user_id)
                    .single()
                    .execute()
                )

                if not profile.data:
                    print(f"[SUBSCRIPTION_UPDATED] Profile not found for user {user_id}")
                else:
                    old_plan = profile.data.get("plan_type", "free")
                    current_credits = profile.data.get("credits_remaining", 0)

                    # Get new plan from subscription items
                    items = subscription.get("items", {}).get("data", [])
                    if items:
                        new_price_id = items[0].get("price", {}).get("id")
                        new_plan = PRICE_TO_PLAN.get(new_price_id, old_plan)

                        if new_plan and new_plan != old_plan:
                            old_plan_credits = CREDITS_MAP.get(old_plan, 0)
                            new_plan_credits = CREDITS_MAP.get(new_plan, 0)
                            
                            # Check if this upgrade was handled by our API (has bonus_credits in metadata)
                            # If so, skip credit allocation here to prevent double-granting
                            if sub_metadata.get("bonus_credits"):
                                print(
                                    f"[SUBSCRIPTION_UPDATED] Upgrade via API detected (bonus_credits={sub_metadata.get('bonus_credits')}), "
                                    f"skipping webhook credit allocation"
                                )
                                # Still update plan_type if needed (shouldn't be, but just in case)
                                if profile.data.get("plan_type") != new_plan:
                                    supabase.table("profiles").update({
                                        "plan_type": new_plan,
                                    }).eq("id", user_id).execute()

                            # UPGRADE: New plan has more credits than old plan (Dashboard-initiated)
                            elif new_plan_credits > old_plan_credits:
                                # Update plan_type AND credits for Dashboard-initiated upgrades
                                res_update = (
                                    supabase.table("profiles")
                                    .update({
                                        "plan_type": new_plan,
                                        "credits_remaining": new_plan_credits,
                                    })
                                    .eq("id", user_id)
                                    .execute()
                                )
                                log_db("upgrade_plan_with_credits", res_update)
                                
                                # Log to ledger
                                supabase.table("ledger").insert({
                                    "user_id": user_id,
                                    "change": new_plan_credits,
                                    "amount": 0,
                                    "reason": f"plan upgrade - {old_plan} to {new_plan} (via Stripe Dashboard)",
                                    "ts": datetime.utcnow().isoformat(),
                                }).execute()

                                print(
                                    f"[UPGRADE] user={user_id}, {old_plan}→{new_plan}, "
                                    f"credits: {current_credits}→{new_plan_credits} (Dashboard)"
                                )

                            # DOWNGRADE: New plan has fewer credits than old plan
                            elif new_plan_credits < old_plan_credits:
                                # Don't change credits yet - downgrade will be applied at next invoice.paid
                                # This happens when user downgrades with proration_behavior='none'
                                print(
                                    f"[DOWNGRADE_SCHEDULED] user={user_id}, {old_plan}→{new_plan}, "
                                    f"will take effect at period end, keeping {current_credits} credits for now"
                                )
                                # Optionally update plan_type here if you want to show the pending change
                                # For now, we'll wait until the next invoice.paid to apply the downgrade

                            else:
                                print(f"[SUBSCRIPTION_UPDATED] No credit change for user {user_id}")

            except Exception as exc:
                print(f"[SUBSCRIPTION_UPDATED] Error processing subscription update: {exc}")
        else:
            print(f"[SUBSCRIPTION_UPDATED] Could not find user_id for subscription {subscription_id}")

    # Handle subscription cancellation/deletion
    elif event_type == "customer.subscription.deleted":
        subscription_id = obj.get("id")
        sub_metadata = obj.get("metadata", {})
        user_id = sub_metadata.get("user_id")

        if not user_id and customer_id:
            user_id = _find_user_by_customer(customer_id)

        if user_id:
            linked_user_id = user_id
            try:
                # Check why the subscription was deleted
                # cancellation_details.reason can be: cancellation_requested, payment_disputed, payment_failed
                cancellation_details = obj.get("cancellation_details", {}) or {}
                cancellation_reason = cancellation_details.get("reason", "unknown")
                canceled_at = obj.get("canceled_at")

                print(f"[SUBSCRIPTION_DELETED] user={user_id}, reason={cancellation_reason}")

                # Get current profile to check if already flagged for fraud
                profile_res = (
                    supabase.table("profiles")
                    .select("credits_remaining, addon_credits, fraud_flag, subscription_status")
                    .eq("id", user_id)
                    .single()
                    .execute()
                )

                if not profile_res.data:
                    print(f"[CANCEL] Profile not found for user {user_id}")
                    _mark_event_processed(event_id, event_type)
                    return {"status": "profile_not_found", "event_id": event_id}

                profile = profile_res.data
                current_credits = profile.get("credits_remaining", 0)
                current_addon = profile.get("addon_credits", 0)
                fraud_flag = profile.get("fraud_flag")
                current_status = profile.get("subscription_status")

                # If already revoked due to fraud, skip (credits already taken)
                if current_status == "revoked" or fraud_flag:
                    print(f"[CANCEL] User {user_id} already revoked (fraud_flag={fraud_flag}), skipping")
                    _mark_event_processed(event_id, event_type)
                    return {"status": "already_revoked", "event_id": event_id}

                # Check if this was a Stripe-initiated deletion due to payment issues
                # These reasons indicate the user didn't properly cancel via the app
                stripe_initiated_reasons = {"payment_disputed", "payment_failed"}

                if cancellation_reason in stripe_initiated_reasons:
                    # Stripe force-canceled due to payment issues - revoke ALL credits
                    print(f"[CANCEL] ⚠️  Stripe-initiated cancellation ({cancellation_reason}) - REVOKING ALL CREDITS")

                    revoke_result = _revoke_all_credits_for_fraud(
                        user_id=user_id,
                        reason=f"Subscription deleted by Stripe - reason={cancellation_reason}",
                        event_type=event_type,
                        event_id=event_id,
                        amount=0,
                        cancel_subscription=False,  # Already deleted
                    )

                    _mark_event_processed(event_id, event_type)
                    return {"status": "credits_revoked_stripe_cancel", "event_id": event_id, "details": revoke_result}

                # Normal cancellation (user canceled via app or period ended naturally)
                # Reset monthly credits to 0, preserve addon credits
                res_update = (
                    supabase.table("profiles")
                    .update({
                        "subscription_status": "canceled",
                        "plan_type": "free",
                        "credits_remaining": 0,  # Reset monthly credits - subscription ended
                        "pending_downgrade": None,  # Clear any pending downgrade
                        # addon_credits preserved - user paid separately for these
                    })
                    .eq("id", user_id)
                    .execute()
                )
                log_db("cancel_subscription", res_update)

                # Log to ledger
                supabase.table("ledger").insert({
                    "user_id": user_id,
                    "change": -current_credits if current_credits > 0 else 0,
                    "amount": 0,
                    "reason": f"Subscription ended (reason={cancellation_reason}) - monthly credits reset, addon preserved",
                    "ts": datetime.utcnow().isoformat(),
                }).execute()

                print(f"[CANCELED] user={user_id}, subscription={subscription_id}, reason={cancellation_reason}, reset monthly: {current_credits}→0, kept addon: {current_addon}")

            except Exception as exc:
                print(f"[CANCEL] Error updating canceled subscription: {exc}")
        else:
            print(f"[CANCEL] Could not find user_id for subscription {subscription_id}, skipping")

    if event_type in STRIPE_SYNC_EVENTS and customer_id:
        sync_result = sync_stripe_customer(customer_id, user_id=linked_user_id)
        print(f"[SYNC] {sync_result}")

    # Mark event as processed to prevent duplicate processing
    _mark_event_processed(event_id, event_type)

    print("========== END WEBHOOK ==========")
    return {"status": "success"}



